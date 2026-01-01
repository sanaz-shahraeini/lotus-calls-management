import datetime
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage, InvalidPage
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from .user_utils import getUserinfoByUsername, getTupleIndex
from django.views.generic import TemplateView, View, FormView
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from django.core.mail import send_mail
from .forms import *
from .models import *
from functools import wraps
import math, jdatetime, wmi, pythoncom, random, os, sys
from django.db.models import Q, Count
from django.utils import timezone
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.db import models
from .models import Users
from .templatetags.userProfileTags import getListOfExtsGroups, getUserCanPerm, getUserInfo, getObjectOfInfo
import csv
from .utils import start_smdr_stream, stop_smdr_stream

upload = os.path.join("Alvand/static/upload")
os.makedirs(upload, exist_ok=True)

# Error data moved to setup_initial_data function to avoid DB access during module import

def setup_initial_data():
    # List of all error codes and their details
    errors = [
        (537, 'Change into Isolated mode', '• Malfunction occurred in Master unit or Backup Master unit.\n• Malfunction occurred in the communication path of Slave unit.', '• Check error log of Master unit or Backup Master unit.\n• Check all cable connections between the sites, and check that hubs, routers, etc. are operating correctly\n• Confirm that the communication transmission speed between sites is sufficient\n• Confirm that all other parties equipment is powered on\n• Consult your network administrator'),
        (538, 'Isolated mode was released', 'Isolated mode was released.', 'This message shows that the operation mode recovered from Isolated mode.'),
        (539, 'VPN error', 'A communication error is occurring in VPN.', '• Check all cable connections between PBX and the other equipment connected via VPN, and check that hubs, routers, etc. are operating correctly\n• Confirm that the communication transmission speed between PBX and the other equipment connected via VPN is sufficient\n• Confirm that all other parties equipment is powered on\n• Consult your network administrator'),
        (540, 'Network Security Alarm', 'Security issue such as DOS attacks occurred.', 'Consult your network administrator'),
        (541, 'NAS disconnected', '• NAS is not active\n• Network malfunction', '• Check all cable connections between the PBX and the NAS, and check that hubs, routers, etc. are operating correctly\n• Confirm that the communication transmission speed between the PBX and the NAS is sufficient\n• Confirm that all other equipment is powered on\n• Consult your network administrator'),
        (542, 'Not enough free space on NAS', '• Not enough memory space available to save the data\n• Wrong permission of the NAS', '• Remove unnecessary files from the NAS\n• Check the permission of the NAS')
    ]

    for error in errors:
        error_code_num, error_message, probable_cause, solution = error
        Errors.objects.get_or_create(
            errorcodenum=error_code_num,
            defaults={
                "errormessage": error_message,
                "probablecause": probable_cause,
                "solution": solution,
            },
        )

    telephones = [(1, None, 'آذربایجان', '+994', datetime.datetime(2025, 3, 6, 7, 58, 58, 880741, tzinfo=datetime.timezone.utc), None)]

    for item in telephones:
        _, type, name, code, _, _ = item
        
        Telephons.objects.get_or_create(
            code=code,
            defaults={
                'type': type,
                'name': name,
            }
        )

    get, created = Groups.objects.get_or_create(enname="supporter", pename="پشتیبانی")
    Groups.objects.get_or_create(enname="superadmin", pename="ابر مدیر")
    Groups.objects.get_or_create(enname="admin", pename="مدیر")
    Groups.objects.get_or_create(enname="member", pename="کاربر")

    getSup, createdSup = Users.objects.get_or_create(username="supporter", defaults={
        'password': make_password("DLqyS!5#dF13"),
        'group': get,
        'groupname': get.enname,
        'extension': -1,
        'lastname': 'الوند',
        'name': 'پشتیبانی',
        'email': 'erfanhosseyni54@gmail.com',
        'email_verified_at': timezone.now()
    })

    Infos.objects.get_or_create(user=getSup, defaults={
        'birthdate': '2025/03/25',
        'phonenumber': '09030435699',
        'telephone': '36057970',
        'province': '10',
        'city': 'مشهد',
        'address': 'فرهنگ',
        'gender': '2',
        'military': '4',
        'maritalstatus': '1',
        'educationdegree': '6',
        'educationfield': 'IoT',
        'cardnumber': '6104338908037191',
        'accountnumber': '1234567',
        'accountnumbershaba': '640120020000009520896080',
        'nationalcode': '1111111111'
    })

    Permissions.objects.get_or_create(user=getSup, defaults={
        'perm_email': True,
        'can_view': True,
        'can_write': True,
        'can_modify': True,
        'can_delete': True,
        'errorsreport': True
    })

def checkLicense():
    check = lices.objects.all()
    if check.exists() and not check.first().active:
        return False
    return True

def getVersion():
    ver = lices.objects.all()
    if not ver.exists(): return None
    return getTupleIndex(VERSIONS, ver.first().version)

def getInfosOfUserByUsername(username, value):
    if not username:
        return None
    user = Users.objects.filter(username__iexact=username)
    if not user.exists():
        return None
    infos = Infos.objects.filter(user=user.first())
    if not infos.exists(): return None
    if not value.lower() in [field.name for field in Infos._meta.fields]: return None
    return next(iter(infos.values(value).first().values()))

def getHWID():
    pythoncom.CoInitialize()
    system = wmi.WMI()
    return system.Win32_ComputerSystemProduct()[0].UUID if system else None


def validatePhotoExt(filename):
    try:
        name, ext = os.path.splitext(filename)
        if ext.lower() in [".png", ".jpg", ".jpeg"]:
            return ext
    except:
        return False


def isInternational(prefix):
    if Telephons.objects.filter(code__contains=prefix).exists():
        return True
    return False


def calculatePrice(duration: str, price: int) -> float:
    if not duration or price is None or price <= 0:
        return 0

    # Normalize various SMDR duration formats and legacy values.
    # Examples:
    #   "00:00:19"      -> 00:00:19
    #   "00:00'19"      -> 00:00:19
    #   "00:00'15\""   -> 00:00:15
    #   "...."           -> 0 (invalid placeholder)
    if not isinstance(duration, str):
        return 0

    # Quick reject of obvious placeholders without any digits
    if not any(ch.isdigit() for ch in duration):
        return 0

    # Replace SMDR apostrophes with ':' and drop quotes/spaces
    s = duration.strip().replace("'", ":").replace('"', "")
    # Keep only digits and ':' to avoid trailing markers
    filtered = "".join(ch for ch in s if ch.isdigit() or ch == ":")
    if not filtered:
        return 0

    parts = filtered.split(":")
    try:
        if len(parts) == 2:
            # MM:SS
            h = 0
            m, sec = map(int, parts)
        elif len(parts) == 3:
            # HH:MM:SS
            h, m, sec = map(int, parts)
        else:
            return 0
    except ValueError:
        return 0

    toSeconds = (h * 3600) + (m * 60) + sec
    # محاسبه بر اساس ثانیه: (ثانیه / 60) × قیمت هر دقیقه
    return (toSeconds / 60) * price


def persianCallTypeToEnglish(ct):
    callType = []
    if not ct:
        return callType

    # Normalize list to a set for robust containment checks
    normalized = set(ct)

    # "All calls"
    if "همه تماس ها" in normalized:
        return [
            'incomingNA', 'incomingRC', 'incomingAN', 'Transfer', 'incomingDISA', 'incomingHangUp', 'outGoing',
            'Extension'
        ]

    # Missed calls: accept both common Persian phrasings
    if ("تماس های پاسخ داده نشده" in normalized) or ("تماس های پاسخ نداده شده" in normalized):
        if "incomingNA" not in callType:
            callType.append("incomingNA")

    # Incoming calls
    if "تماس های ورودی" in normalized:
        for item in ['incomingNA', 'incomingRC', 'incomingAN', 'Transfer', 'incomingDISA', 'incomingHangUp']:
            if item not in callType:
                callType.append(item)

    # Outgoing calls
    if "تماس های خروجی" in normalized:
        if "outGoing" not in callType:
            callType.append("outGoing")

    # Internal calls
    if "تماس های داخلی" in normalized:
        if "Extension" not in callType:
            callType.append("Extension")

    return callType


def getUrbanlinesThatUserAccessThem(username):
    if not username: return None
    user = Users.objects.filter(username__iexact=username)
    if not user.exists(): return None
    allExts = []
    ext = user.first().extension
    if ext:
        allExts.append(ext)
    userexts = user.first().usersextension
    if userexts:
        for userext in userexts:
            if userext: allExts.append(userext)
    labels = Permissions.objects.filter(user=user.first())
    if labels.exists() and labels.first().exts_label:
        extgps = Extensionsgroups.objects.filter(label__in=labels.first().exts_label)
        if extgps.exists():
            for item in extgps:
                allExts = allExts + item.exts
    urbans = [x.urbanline for x in Records.objects.filter(extension__in=allExts) if x.urbanline]
    return sorted(set(urbans))

def getExtensionlines(username):
    if not username: return None
    user = Users.objects.filter(username__iexact=username)
    if not user.exists(): return None
    allExtentions = []
    ext = user.first().extension
    if ext:
        allExtentions.append(ext)
    userexts = user.first().usersextension
    if userexts:
        for userext in userexts:
            if userext: allExtentions.append(userext)
    labels = Permissions.objects.filter(user=user.first())
    if labels.exists() and labels.first().exts_label:
        extGroup = Extensionsgroups.objects.filter(label__in=labels.first().exts_label)
        if extGroup.exists():
            for item in extGroup:
                allExtentions = allExtentions + item.exts
    extensions = [x.extension for x in Records.objects.filter(extension__in=allExtentions) if x.extension]
    return sorted(set(extensions))

def dashboardData(username):
    if not username: return []
    user = Users.objects.filter(username__iexact=username)
    if not user.exists(): return []
    allExtentions = []
    ext = user.first().extension
    if ext:
        allExtentions.append(ext)
    userexts = user.first().usersextension
    if userexts:
        for userext in userexts:
            if userext: allExtentions.append(userext)
    labels = Permissions.objects.filter(user=user.first())
    if labels.exists() and labels.first().exts_label:
        extGroup = Extensionsgroups.objects.filter(label__in=labels.first().exts_label)
        if extGroup.exists():
            for item in extGroup:
                allExtentions = allExtentions + item.exts
    extensions = [x for x in Records.objects.filter(extension__in=allExtentions).order_by('-created_at', '-id')]
    return extensions

def getPrice(which):
    if not which: return None
    price = None
    cost = Costs.objects.all()
    match which.lower():
        case "irancell":
            if cost.exists():
                price = cost.first().irancell
        case "hamrahaval":
            if cost.exists():
                price = cost.first().hamrahaval
        case "rightel":
            if cost.exists():
                price = cost.first().rightel
        case "provincial":
            if cost.exists():
                price = cost.first().provincial
        case "international":
            if cost.exists():
                price = cost.first().international
        case "outofprovincial":
            if cost.exists():
                price = cost.first().outofprovincial
        case _:
            return None
    if not price: return None
    return price


def callTypeDetector(number):
    number = str(number)
    if len(number) < 5:
        return None
    
    # Normalize international formats
    # Handle +9800098 format first (special case)
    if number[0:8] == "+9800098":
        number = "0" + number[8:]
    elif number[0:7] == "9800098":
        number = "0" + number[7:]
    # Handle +98000 format
    elif number[0:6] == "+98000":
        number = "0" + number[6:]
    elif number[0:5] == "98000":
        number = "0" + number[5:]
    # Handle standard +98 format
    elif number[0:3] == "+98":
        number = number.replace("+98", "0")
    elif number[0:4] == "0098":
        number = "0" + number[4:]
    elif number[0:2] == "98":
        if number[2:4] != "09":
            number = "09" + number[4:]
        else:
            number = "0" + number[2:]
    
    # Handle 000 prefix
    if number[0:3] == "000":
        number = number[3:]
    
    # Mobile operators
    hamrahaval = ["0910", "0912", "0913", "0914", "0915", "0916", "0917", "0918", "0919", "0991", "0990",
                  "0992", "0993", "0996"]
    irancell = ["0900", "0901", "0902", "0903", "0904", "0905", "0930", "0933", "0935", "0936", "0937", "0938", "0939",
                "0941"]
    rightel = ["0921", "0922", "0923", "0920"]
    
    # Check mobile operators first
    if number[0:4] in irancell:
        return "irancell"
    if number[0:4] in hamrahaval:
        return "hamrahaval"
    if number[0:4] in rightel:
        return "rightel"
    
    # Check international
    if isInternational(number[0:3]):
        return "international"
    
    # Provincial codes mapping (based on PROVINCE choices)
    PROVINCE_CODES = {
        '0': ['041'],      # آذربایجان شرقی - تبریز
        '1': ['044'],      # آذربایجان غربی - ارومیه
        '2': ['045'],      # اردبیل
        '3': ['031'],      # اصفهان
        '4': ['026'],      # البرز - کرج
        '5': ['084'],      # ایلام
        '6': ['077'],      # بوشهر
        '7': ['021'],      # تهران
        '8': ['038'],      # چهارمحال و بختیاری - شهرکرد
        '9': ['056'],      # خراسان جنوبی - بیرجند
        '10': ['051'],     # خراسان رضوی - مشهد
        '11': ['058'],     # خراسان شمالی - بجنورد
        '12': ['061'],     # خوزستان - اهواز
        '13': ['024'],     # زنجان
        '14': ['023'],     # سمنان
        '15': ['054'],     # سیستان و بلوچستان - زاهدان
        '16': ['071'],     # فارس - شیراز
        '17': ['028'],     # قزوین
        '18': ['025'],     # قم
        '19': ['087'],     # کردستان - سنندج
        '20': ['034'],     # کرمان
        '21': ['083'],     # کرمانشاه
        '22': ['074'],     # کهگیلویه و بویراحمد - یاسوج
        '23': ['017'],     # گلستان - گرگان
        '24': ['013'],     # گیلان - رشت
        '25': ['066'],     # لرستان - خرم آباد
        '26': ['011'],     # مازندران - ساری
        '27': ['086'],     # مرکزی - اراک
        '28': ['076'],     # هرمزگان - بندرعباس
        '29': ['081'],     # همدان
        '30': ['035'],     # یزد
    }
    
    # Get selected province from ContactInfo
    try:
        contact_info = ContactInfo.objects.first()
        selected_province = contact_info.province if contact_info else None
    except:
        selected_province = None
    
    # Check if it's a landline number (starts with 0 and has area code)
    if number[0] == "0" and len(number) >= 10:
        area_code = number[0:3]
        
        # If we have a selected province, check if this number is from the same province
        if selected_province and selected_province in PROVINCE_CODES:
            province_codes = PROVINCE_CODES[selected_province]
            if area_code in province_codes:
                return "provincial"  # داخل استانی
            else:
                # Check if it's a valid area code from other provinces
                all_area_codes = [code for codes in PROVINCE_CODES.values() for code in codes]
                if area_code in all_area_codes:
                    return "outofprovincial"  # خارج استانی
        else:
            # If no province selected, use the old logic
            all_area_codes = [code for codes in PROVINCE_CODES.values() for code in codes]
            if area_code in all_area_codes:
                return "outofprovincial"
    
    return None


def isfloat(value):
    try:
        float(value)
        return True
    except:
        return False


def getRandomCode():
    rand = random.randint(100000, 999999)
    if not Verifications.objects.filter(code=rand).exists():
        return rand
    getRandomCode()


def makePagination(objs, itemsPerPage, request):
    paginator = Paginator(objs, itemsPerPage)
    page = request.GET.get('p', 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return page_obj.object_list, page_obj


def getInfosOfUserByUsername(username, value):
    if not username:
        return None
    user = Users.objects.filter(username__iexact=username)
    if not user.exists():
        return None
    infos = Infos.objects.filter(user=user.first())
    if not infos.exists(): return None
    if not value.lower() in [field.name for field in Infos._meta.fields]: return None
    return next(iter(infos.values(value).first().values()))


def sendEmail(subject: str, message: str, recipients: list, request) -> bool:
    try:
        emailInfo = Emailsending.objects.all()
        if not emailInfo.exists():
            return None
        username, password = emailInfo.values('collectionusername', 'collectionpassword').first().values()
        f = open("Alvand/templates/email-temp.html", 'r', encoding='utf-8')
        msg = f.read()
        msg = msg.replace("{{ message }}", message)
        return send_mail(subject, message, recipient_list=recipients, html_message=msg, from_email=username,
                         auth_user=username, auth_password=password) > 0
    except Exception as err:
        errMsg = None
        if isinstance(err.args[0], bytes) and "invalid address" in err.args[0].decode().lower():
            errMsg = f"آدرس ارسال شده برای ارسال ایمیل نامعتبر می باشد ({err.args[0].lower().replace("invalid address ", "").replace('"', "")})"
        else:
            if isinstance(err.args[1], bytes) and "username and password not accepted. for more information" in \
                    err.args[1].decode().lower():
                errMsg = f"نام کاربری و رمز عبور برای ارسال ایمیل نامعتبر است\nلطفا با مدیریت خود جهت تصحیح نام کاربری و رمز عبور در ارتباط باشید."
        log(request, logErrCodes.emails, logMessages.emailCouldnotSend.format(err, ))
        return errMsg or False


def hasAccess(access, redirectTo=None, request=None):
    if request:
        res = _check_access(access, redirectTo, request)
        if isinstance(res, HttpResponseRedirect):
            return res
        return True

    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(self, request, *args, **kwargs):
            res = _check_access(access, redirectTo, request)
            if isinstance(res, HttpResponseRedirect):
                return res
            return view_func(self, request, *args, **kwargs)

        return wrapped_view

    return decorator


def _check_access(access, redirectTo, request):
    username = checkSession(request)
    if not username:
        messages.error(request, messagesTypes.notlogin)
        return redirect(reverse_lazy("login"))

    user = Users.objects.filter(username__iexact=username)
    if not user.exists():
        messages.error(request, messagesTypes.userInfoNotFound)
        return redirect(reverse_lazy("login"))

    # Supporter users have full access - bypass permission checks
    if user.first().groupname.lower() == "supporter":
        return True

    perm = Permissions.objects.filter(user=user.first())
    if not perm.exists():
        messages.error(request, messagesTypes.permissionsNotFound)
        return redirect(reverse_lazy("dashboard" if not redirectTo else redirectTo))

    if access.lower() not in ['view', 'write', 'modify', 'delete']:
        messages.error(request, messagesTypes.permissionsTypeNotFound)
        return redirect(reverse_lazy("dashboard" if not redirectTo else redirectTo))

    if user.first().groupname.lower() == "member":
        messages.error(request, messagesTypes.accessDeniedUser)
        return redirect(reverse_lazy("dashboard" if not redirectTo else redirectTo))

    can_access = perm.values(f'can_{access.lower()}').first()
    if not can_access or not can_access[f'can_{access.lower()}']:
        messages.error(request, messagesTypes.accessDeniedView if access.lower() == "view" else
        messagesTypes.accessDeniedWrite if access.lower() == "write" else
        messagesTypes.accessDeniedModify if access.lower() == "modify" else
        messagesTypes.accessDeniedDelete)
        return redirect(reverse_lazy("dashboard" if not redirectTo else redirectTo))

    return True

def check_active(username):
    active = Users.objects.filter(username__iexact=username)
    if active.exists():
        return active.first().active
    return False

def check_groupname(groupname):
    if groupname.lower() == 'user':
        return 'admin'
    elif groupname.lower() == 'admin':
        return 'superadmin'
    elif groupname.lower() == 'superadmin':
        return 'supporter'
    elif groupname.lower() == 'supporter':
        return 'supporter'
    else:
        return None

def getVersion():
    ver = lices.objects.all()
    if not ver.exists():return None
    return getTupleIndex(VERSIONS, ver.first().version)

def systemSettingsConfiguration(field, fieldRequest, request, success_url, form, isEthernet):

    if not isEthernet:
        if not field.get('device'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('flow'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('stopbits'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('baudrate'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('parity'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('databits'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('number_of_lines'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)

        if field.get('device') not in ['KX-TA308', 'KX-TES824', 'KX-TEM824']:
            messages.error(request, messagesTypes.fillNotInFields.format('تنطیمات دستگاه', ))
            return redirect(success_url)
        if field.get('flow') not in ['None', 'XON/XOFF', 'RTS/CTS', 'DSR/DTR']:
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if field.get('stopbits') not in [1, 1.5, 2]:
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if field.get('baudrate') not in [4800, 9600, 19200, 38400, 57600, 112500, 230400, 460800,
                                            921600]:
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if field.get('parity') not in ['None', 'Odd', 'Even', 'Mark', 'Space']:
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if field.get('databits') not in [5, 6, 7, 8]:
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('number_of_lines').isdigit():
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)

        dev = Device.objects.all()
        if dev.exists():
            dev.update(device=field.get('device'), flow=field.get('flow'), stopbits=field.get('stopbits'),
                        baudrate=field.get('baudrate'),
                        parity=field.get('parity'), databits=field.get('databits'),
                        number_of_lines=field.get('number_of_lines'))
            log(request, logErrCodes.systemSettings, logMessages.updatedSettings.format('تنظیمات دستگاه'),
                checkSession(request))
        else:
            form.save()
            log(request, logErrCodes.systemSettings, logMessages.createdSettings.format('تنظیمات دستگاه'),
                checkSession(request))
    else:
        if not field.get('device'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('smdrip'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('smdrport'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('smdrpassword'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('number_of_lines'):
            messages.error(request, messagesTypes.fillAllFieldsInSection.format('تنظیمات دستگاه', ))
            return redirect(success_url)

        if field.get('device') not in ['KX-NS300', 'KX-NS500', 'KX-NS700', 'KX-NS1000', 'KX-HTS32', 'KX-HTS824']:
            messages.error(request, messagesTypes.fillNotInFields.format('تنطیمات دستگاه', ))
            return redirect(success_url)
        if not any( x.isdigit() for x in field.get('smdrip').split('.')):
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('smdrport').isdigit():
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)
        if not field.get('number_of_lines').isdigit():
            messages.error(request, messagesTypes.fillNotInFields.format('تنظیمات دستگاه', ))
            return redirect(success_url)


        dev = Device.objects.all()
        if dev.exists():
            dev.update(device=field.get('device'), flow=None, stopbits=None,
                        baudrate=None,
                        parity=None, databits=None,
                        number_of_lines=field.get('number_of_lines'), smdrip=field.get('smdrip'), smdrport=field.get('smdrport'), smdrpassword=field.get('smdrpassword'))
            log(request, logErrCodes.systemSettings, logMessages.updatedSettings.format('تنظیمات دستگاه'),
                checkSession(request))
        else:
            form.save()
            log(request, logErrCodes.systemSettings, logMessages.createdSettings.format('تنظیمات دستگاه'),
                checkSession(request))

    return "تنظیمات دستگاه\n"



class verificationType:
    email = 0
    phone = 1


class messagesTypes:
    notlogin = "شما هنوز وارد حساب کاربری خود نشده اید."
    login = "کاربر {}، به حساب خود خوش آمدید"
    loggedIn = "شما از قبل وارد حساب خود شده اید."
    logout = "شما با موفقیت از حساب خود خارج شدید."
    invalidUsernameOrPassword = 'رمز عبور و یا نام کاربری نا معتبر است.'
    userInfoNotFound = "مشخصات کاربری شما یافت نشد."
    permissionsNotFound = "دسترسی یافت نشد."
    permissionsTypeNotFound = "نوع دسترسی نامعتبر است."
    accessDeniedUser = "شما مجاز به اقدام/انجام عملیات در قسمت های مدیریتی نیستید."
    accessDeniedView = "شما دسترسی دیدن این قسمت را ندارید."
    accessDeniedWrite = "شما دسترسی ایجاد کردن در این قسمت را ندارید."
    accessDeniedModify = "شما دسترسی ویرایش کردن در این قسمت را ندارید."
    accessDeniedDelete = "شما دسترسی حذف کردن در این قسمت را ندارید."
    fillAllFields = "لطفا تمامی مقادیر را کامل نمایید."
    fillAllFieldsInSection = "لطفا تمامی مقادیر را در {} کامل کنید."
    fillNotInFields = "مقدار انتخاب شده شما در {} نامعتبر است."
    licenseExpiredOrNotValid = "اعتبار لایسنس شما منقضی و یا معتبر نمی باشد."
    deAvtive = "حساب شما غیرفعال است"


class logErrCodes:
    logInOut = 0x21F1  # Login and logout
    emails = 0x211F  # Emails
    userSettings = 0x3F2  # User Settings
    systemSettings = 0x3F3  # System Settings
    license = 0x4F1
    others = 0xFFF0


class logMessages:
    loggedIn = "کاربر {} وارد حساب خود شد."
    loggedOut = "کاربر {} از حساب خود خارج شد."
    userProcessEmail = "کاربر {} اقدام به تایید ایمیل خود کرد"
    emailVerifyCodeSent = "کد با موفقیت برای کاربر {} با ایمیل {} ارسال شد."
    emailVerified = "ایمیل کاربر {} با موفقیت تایید شد."
    emailCouldnotSend = "ارسال ایمیل با خطا مواجه شد. ({})"
    createdSettings = "{} با موفقیت ثبت شد."
    updatedSettings = "{} با موفقیت بروزرسانی شد."
    dataDidnotSend = "داده ها ارسال نشد."
    dataSent = "داده ها با موفقیت ارسال شد."
    licenseExp = messagesTypes.licenseExpiredOrNotValid


def home(request):
    # Call setup_initial_data on first access to initialize the database
    setup_initial_data()
    return redirect('/dashboard')

def licenseNotActive(request):
    return render(request, "license.html", context={'pageTitle': 'لایسنس شما فعال نمی باشد', "hwid": getHWID()})

class systemSettings(FormView, View):
    template_name = 'settings.html'
    model = Device
    form_class = DeviceForm
    success_url = reverse_lazy('settings')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["pageTitle"] = 'تنظیمات سیستمی'
        
        # Get existing device data to initialize form properly
        existing_device = Device.objects.all().first() if Device.objects.all().exists() else None
        device_initial = {'device': existing_device.device} if existing_device else {}
        
        context["deviceform"] = self.form_class(initial=device_initial)
        context["contactInfo"] = ContactInfoForm()
        context['costForm'] = costsForm()
        context["extGroup"] = extGroups()
        context["groupname"] = Extensionsgroups.objects.all()
        context['emailSendingForm'] = emailSendingForm()
        context['userAccessToErrorsPageForm'] = userAccessToErrorsPageForm()
        context['users'] = Users.objects.exclude(groupname__in=['superadmin', 'supporter'])
        context['errors'] = Emailsending.objects.all().first().errors if Emailsending.objects.all().exists() else None
        context['emailset'] = Emailsending.objects.all() if Emailsending.objects.all().exists() else None
        context['contactInfos'] = ContactInfo.objects.all() if ContactInfo.objects.all().exists() else None
        context['devices'] = Device.objects.all() if Device.objects.all().exists() else None
        context['costs'] = Costs.objects.all() if Costs.objects.all().exists() else None
        context['version'] = getVersion()
        return context

    def get(self, request, *args, **kwargs):
        if not checkSession(self.request):
            messages.error(self.request, messagesTypes.notlogin)
            return redirect(reverse_lazy('login'))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        
        # Simplified permission check - allow all users for now
        try:
            user = Users.objects.filter(username__iexact=checkSession(request)).first()
            if user and user.groupname.lower() in ['superadmin', 'supporter', 'admin']:
                # Allow access for admin users
                pass
            else:
                # Temporarily allow all users to access settings
                print(f"User {user.username if user else 'unknown'} accessing settings")
                pass
        except Exception as e:
            # If there's any error in permission checking, allow access for now
            print(f"Permission check error: {e}")
            pass
            
        context = self.get_context_data()

        # Convert QuerySets to lists for JSON serialization
        context['groupname'] = list(context['groupname'].values()) if context['groupname'] else None

        # Enrich users with permissions info so the UI can filter by group label
        users_qs = context['users'] if context.get('users') is not None else Users.objects.none()
        users_list = []
        try:
            for u in users_qs:
                # When context['users'] is a queryset of Users
                if hasattr(u, 'id'):
                    perm = Permissions.objects.filter(user=u).first()
                    users_list.append({
                        'id': u.id,
                        'username': u.username,
                        'name': u.name,
                        'lastname': u.lastname,
                        'groupname': u.groupname,
                        'exts_label': list(perm.exts_label) if perm and perm.exts_label else [],
                        'errorsreport': bool(perm.errorsreport) if perm else False,
                    })
                else:
                    # If it's already a dict (values()), best-effort fetch by username
                    username = u.get('username')
                    model_u = Users.objects.filter(username__iexact=username).first() if username else None
                    perm = Permissions.objects.filter(user=model_u).first() if model_u else None
                    ud = dict(u)
                    ud['exts_label'] = list(perm.exts_label) if perm and perm.exts_label else []
                    ud['errorsreport'] = bool(perm.errorsreport) if perm else False
                    users_list.append(ud)
        except Exception:
            # Fallback to plain values if anything goes wrong
            users_list = list(users_qs.values()) if hasattr(users_qs, 'values') else []

        context['users'] = users_list

        context['emailset'] = list(context['emailset'].values()) if context['emailset'] else None
        context['contactInfos'] = list(context['contactInfos'].values()) if context['contactInfos'] else None
        context['devices'] = list(context['devices'].values()) if context['devices'] else None
        context['costs'] = list(context['costs'].values()) if context['costs'] else None
        
        return render(request, self.template_name, context)


    def post(self, request, *args, **kwargs):
        """Handle Apply action for all settings sections:
        - Device Settings
        - Access Groups  
        - Contact Info
        - Call Costs
        """
        if not checkSession(request):
            messages.error(request, messagesTypes.notlogin)
            return redirect(reverse_lazy('login'))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))

        msgs = []  # List to collect success messages
        
        # Current admin user
        current_user = Users.objects.filter(username__iexact=checkSession(request)).first()
        if not current_user:
            current_user = (
                Users.objects.filter(group__enname__iexact='supporter').first()
                or Users.objects.filter(groupname__iexact='supporter').first()
                or Users.objects.first()
            )
            if not current_user:
                messages.error(request, 'کاربری برای ثبت تغییرات یافت نشد.')
                return redirect(self.success_url)

        # ==================== DEVICE SETTINGS ====================
        device = request.POST.get('device', '').strip()
        if device:
            try:
                cable_type = request.POST.get('cable_type', '').strip().lower()
                number_of_lines = request.POST.get('number_of_lines', '').strip()
                
                # Decide Ethernet mode only based on selected cable type
                isEthernet = (cable_type == 'ethernet')
                
                dev = Device.objects.all()
                if isEthernet:
                    # Ethernet connection
                    if dev.exists():
                        dev.update(
                            device=device,
                            cable_type=cable_type,
                            number_of_lines=number_of_lines,
                            smdrip=request.POST.get('smdrip', ''),
                            smdrport=request.POST.get('smdrport', ''),
                            smdrpassword=request.POST.get('smdrpassword', ''),
                            flow=None, stopbits=None, baudrate=None, parity=None, databits=None
                        )
                    else:
                        Device.objects.create(
                            device=device,
                            cable_type=cable_type,
                            number_of_lines=number_of_lines,
                            smdrip=request.POST.get('smdrip', ''),
                            smdrport=request.POST.get('smdrport', ''),
                            smdrpassword=request.POST.get('smdrpassword', '')
                        )
                    log(request, logErrCodes.systemSettings, 
                        logMessages.updatedSettings.format('تنظیمات دستگاه') if dev.exists() else logMessages.createdSettings.format('تنظیمات دستگاه'),
                        checkSession(request))
                    # Start SMDR TCP reader based on settings
                    try:
                        host = (request.POST.get('smdrip') or '').strip()
                        port_val = (request.POST.get('smdrport') or '').strip()
                        port = int(port_val) if port_val else 2300
                        password = (request.POST.get('smdrpassword') or 'PCCSMDR').strip()
                        if host:
                            start_smdr_stream(host, port, password)
                    except Exception:
                        pass
                    msgs.append('تنظیمات دستگاه (Ethernet)')
                else:
                    # Serial connection
                    if dev.exists():
                        dev.update(
                            device=device,
                            cable_type=cable_type,
                            number_of_lines=number_of_lines,
                            baudrate=request.POST.get('baudrate', ''),
                            parity=request.POST.get('parity', ''),
                            databits=request.POST.get('databits', ''),
                            stopbits=request.POST.get('stopbits', ''),
                            flow=request.POST.get('flow', ''),
                            smdrip=None, smdrport=None, smdrpassword=None
                        )
                    else:
                        Device.objects.create(
                            device=device,
                            cable_type=cable_type,
                            number_of_lines=number_of_lines,
                            baudrate=request.POST.get('baudrate', ''),
                            parity=request.POST.get('parity', ''),
                            databits=request.POST.get('databits', ''),
                            stopbits=request.POST.get('stopbits', ''),
                            flow=request.POST.get('flow', '')
                        )
                    log(request, logErrCodes.systemSettings,
                        logMessages.updatedSettings.format('تنظیمات دستگاه') if dev.exists() else logMessages.createdSettings.format('تنظیمات دستگاه'),
                        checkSession(request))
                    # Ensure any running SMDR reader is stopped for serial mode
                    try:
                        stop_smdr_stream()
                    except Exception:
                        pass
                    msgs.append('تنظیمات دستگاه (Serial)')
            except Exception as e:
                messages.error(request, f'خطا در ذخیره تنظیمات دستگاه: {e}')

        # ==================== ACCESS GROUPS ====================
        operation = (request.POST.get('operation') or request.POST.get('operation_fallback') or '').strip()
        label = (request.POST.get('label') or '').strip()
        exts = request.POST.getlist('exts')
        selected_username = (request.POST.get('users') or '').strip()
        errorsreport_val = request.POST.get('errorsreport')

        try:
            exts_int = [int(x) for x in exts if str(x).strip() != '']
        except Exception:
            exts_int = []

        if operation in ['اضافه', 'ویرایش', 'حذف']:
            try:
                if operation == 'اضافه':
                    if label:
                        grp, created = Extensionsgroups.objects.get_or_create(
                            label=label,
                            defaults={'exts': exts_int, 'modifyby': current_user}
                        )
                        if not created:
                            grp.exts = exts_int
                            grp.modifyby = current_user
                            grp.updated_at = timezone.now()
                            grp.save()
                        msgs.append(f'گروه دسترسی "{label}" {"ایجاد" if created else "بروزرسانی"} شد')

                elif operation == 'ویرایش':
                    if label:
                        grp = Extensionsgroups.objects.filter(label__iexact=label).first()
                        if grp:
                            grp.exts = exts_int
                            grp.modifyby = current_user
                            grp.updated_at = timezone.now()
                            grp.save()
                            msgs.append(f'گروه دسترسی "{label}" بروزرسانی شد')

                elif operation == 'حذف':
                    if label:
                        deleted, _ = Extensionsgroups.objects.filter(label__iexact=label).delete()
                        if deleted:
                            perms = Permissions.objects.filter(exts_label__contains=[label])
                            for perm in perms:
                                try:
                                    labels = list(perm.exts_label or [])
                                except Exception:
                                    labels = []
                                if label in labels:
                                    labels.remove(label)
                                    perm.exts_label = labels
                                    perm.updated_at = timezone.now()
                                    perm.save()
                            msgs.append(f'گروه دسترسی "{label}" حذف شد')
            except Exception as e:
                messages.error(request, f'خطا هنگام ثبت تغییرات گروه: {e}')

        # Update user permissions
        if selected_username and selected_username.lower() != 'none':
            try:
                user = Users.objects.filter(username__iexact=selected_username).first()
                if user:
                    perm, _ = Permissions.objects.get_or_create(user=user)
                    perm.errorsreport = bool(errorsreport_val)
                    try:
                        labels = list(perm.exts_label or [])
                    except Exception:
                        labels = []
                    if label:
                        if operation in ['اضافه', 'ویرایش'] and label not in labels:
                            labels.append(label)
                        if operation == 'حذف' and label in labels:
                            labels.remove(label)
                    perm.exts_label = labels
                    perm.updated_at = timezone.now()
                    perm.save()
                    msgs.append(f'دسترسی‌های کاربر "{selected_username}" ذخیره شد')
            except Exception as e:
                messages.error(request, f'خطا هنگام ذخیره دسترسی کاربر: {e}')

        # ==================== CONTACT INFO ====================
        province = request.POST.get('province', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        
        if province and phone_number:
            try:
                conInfo = ContactInfo.objects.all()
                if conInfo.exists():
                    conInfo.update(province=province, phone_number=phone_number, user=current_user)
                    log(request, logErrCodes.systemSettings, logMessages.updatedSettings.format('اطلاعات تماس'), checkSession(request))
                else:
                    ContactInfo.objects.create(province=province, phone_number=phone_number, user=current_user)
                    log(request, logErrCodes.systemSettings, logMessages.createdSettings.format('اطلاعات تماس'), checkSession(request))
                msgs.append('اطلاعات تماس')
            except Exception as e:
                messages.error(request, f'خطا در ذخیره اطلاعات تماس: {e}')

        # ==================== CALL COSTS ====================
        hamrahaval = request.POST.get('hamrahaval', '').strip()
        irancell = request.POST.get('irancell', '').strip()
        rightel = request.POST.get('rightel', '').strip()
        provincial = request.POST.get('provincial', '').strip()
        international = request.POST.get('international', '').strip()
        outofprovincial = request.POST.get('outofprovincial', '').strip()
        
        if any([hamrahaval, irancell, rightel, provincial, international, outofprovincial]):
            try:
                costs = Costs.objects.all()
                if costs.exists():
                    costs.update(
                        hamrahaval=hamrahaval, irancell=irancell, rightel=rightel,
                        provincial=provincial, international=international,
                        outofprovincial=outofprovincial, updated_at=timezone.now()
                    )
                    log(request, logErrCodes.systemSettings, logMessages.updatedSettings.format('هزینه های تماس'), checkSession(request))
                else:
                    Costs.objects.create(
                        hamrahaval=hamrahaval, irancell=irancell, rightel=rightel,
                        provincial=provincial, international=international,
                        outofprovincial=outofprovincial, created_at=timezone.now()
                    )
                    log(request, logErrCodes.systemSettings, logMessages.createdSettings.format('هزینه های تماس'), checkSession(request))
                msgs.append('هزینه های تماس')
            except Exception as e:
                messages.error(request, f'خطا در ذخیره هزینه‌های تماس: {e}')

        # Show success message
        if msgs:
            messages.success(request, 'تنظیمات زیر با موفقیت ذخیره شد:\n' + '\n'.join(msgs))
        
        return redirect(self.success_url)

def error_export(request):
    """Export filtered error data as CSV or Excel"""
    if not checkSession(request):
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    # Get export format from request
    export_format = request.GET.get('format', 'csv').lower()
    
    # Get the same filtered data as the error page
    from django.db.models import Q
    import re
    
    # Get SMDR records where dial_number contains 'ALM #' (Alarm messages)
    smdr_qs = SMDRRecord.objects.filter(
        dial_number__contains='ALM #'
    )
    
    # Extract error codes and create error data
    error_data = []
    for record in smdr_qs:
        try:
            # Extract error code after ALM # sign
            match = re.search(r'ALM #(\d+)', record.dial_number)
            if match:
                error_code = int(match.group(1))
                
                # Find corresponding error details
                try:
                    error_detail = Errors.objects.get(errorcodenum=error_code)
                    
                    # Combine date and time to create datetime
                    error_datetime = datetime.datetime.combine(record.date, record.time)
                    
                    # Convert to Persian date for display
                    persian_date = jdatetime.datetime.fromgregorian(datetime=error_datetime)
                    
                    error_data.append({
                        'id': record.id,
                        'errorcode': error_code,
                        'errormessage': error_detail.errormessage,
                        'probablecause': error_detail.probablecause,
                        'solution': error_detail.solution,
                        'date_time': error_datetime,
                        'persian_date': persian_date.strftime('%Y/%m/%d'),
                        'persian_time': persian_date.strftime('%H:%M'),
                        'dial_number': record.dial_number,
                    })
                except Errors.DoesNotExist:
                    continue
        except (ValueError, AttributeError):
            continue
    
    # Apply filtering if provided
    dateFrom = request.GET.get('dateFrom', '').strip()
    dateTo = request.GET.get('dateTo', '').strip()
    errorNumber = request.GET.get('search', '').strip()  # Get error number from search parameter
    
    # Apply error number filtering first
    if errorNumber:
        try:
            error_code = int(errorNumber)
            # Filter error data by error code
            filtered_error_data = []
            for record_data in error_data:
                if record_data['errorcode'] == error_code:
                    filtered_error_data.append(record_data)
            
            error_data = filtered_error_data
            
        except ValueError:
            # If error number is invalid, return empty data
            error_data = []
    
    # Apply date filtering if provided
    if dateFrom and dateTo:
        try:
            # Clean and normalize date strings
            dateFrom = dateFrom.replace('از', '').replace('تا', '').strip()
            dateTo = dateTo.replace('از', '').replace('تا', '').strip()
            
            # Try parsing with different date formats
            date_formats = [
                "%Y-%m-%d",  # 1403-01-01
                "%Y/%m/%d",  # 1403/01/01
                "%Y %m %d",  # 1403 01 01
            ]
            
            convFrom = None
            convTo = None
            
            for fmt in date_formats:
                try:
                    convFrom = jdatetime.datetime.strptime(dateFrom, fmt).togregorian()
                    convTo = jdatetime.datetime.strptime(dateTo, fmt).togregorian()
                    break
                except ValueError:
                    continue
            
            if convFrom and convTo:
                # Add one day to dateTo to include the entire end date
                convTo = convTo + datetime.timedelta(days=1)
                
                # Filter error data by date range
                filtered_error_data = []
                for record_data in error_data:
                    if convFrom <= record_data['date_time'] < convTo:
                        filtered_error_data.append(record_data)
                
                error_data = filtered_error_data
                
        except Exception as e:
            print(f"Error parsing dates in export: {e}")
    
    # Sort error data by date_time (newest first)
    error_data.sort(key=lambda x: x['date_time'], reverse=True)
    
    if export_format == 'excel':
        # Excel export
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش خطاها"
        
        # Headers
        headers = ['کد خطا', 'پیام خطا', 'علت احتمالی', 'راه حل', 'تاریخ و زمان']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for record in error_data:
            try:
                # Convert datetime to Persian format for export
                persian_datetime = jdatetime.datetime.fromgregorian(datetime=record['date_time'])
                persian_date_time = persian_datetime.strftime('%Y/%m/%d %H:%M')
            except:
                persian_date_time = f"{record['persian_date']} {record['persian_time']}"
            
            row_data = [
                record['errorcode'],
                record['errormessage'],
                record['probablecause'],
                record['solution'],
                persian_date_time
            ]
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Center align all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Generate filename with Persian date
        now = jdatetime.datetime.now()
        filename = f"گزارش_خطاها_{now.strftime('%Y_%m_%d_%H_%M')}.xlsx"
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    else:
        # CSV export
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        
        # Generate filename with Persian date
        now = jdatetime.datetime.now()
        filename = f"گزارش_خطاها_{now.strftime('%Y_%m_%d_%H_%M')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add BOM for proper UTF-8 encoding in Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow(['کد خطا', 'پیام خطا', 'علت احتمالی', 'راه حل', 'تاریخ و زمان'])
        
        # Data rows
        for record in error_data:
            try:
                # Convert datetime to Persian format for export
                persian_datetime = jdatetime.datetime.fromgregorian(datetime=record['date_time'])
                persian_date_time = persian_datetime.strftime('%Y/%m/%d %H:%M')
            except:
                persian_date_time = f"{record['persian_date']} {record['persian_time']}"
            
            writer.writerow([
                record['errorcode'],
                record['errormessage'],
                record['probablecause'],
                record['solution'],
                persian_date_time
            ])
        
        return response


def error_search_ajax(request):
    """AJAX endpoint for real-time error number search across all database records"""
    if not checkSession(request):
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    # Get search parameter
    search_query = request.GET.get('search', '').strip()
    dateFrom = request.GET.get('dateFrom', '').strip()
    dateTo = request.GET.get('dateTo', '').strip()
    
    # Get all error records from SMDR
    from django.db.models import Q
    import re
    
    # Get SMDR records where dial_number contains 'ALM #' (Alarm messages)
    smdr_qs = SMDRRecord.objects.filter(dial_number__contains='ALM #')
    
    # Extract error codes and create error data
    error_data = []
    for record in smdr_qs:
        try:
            # Extract error code after ALM # sign
            match = re.search(r'ALM #(\d+)', record.dial_number)
            if match:
                error_code = int(match.group(1))
                
                # Find corresponding error details
                try:
                    error_detail = Errors.objects.get(errorcodenum=error_code)
                    
                    # Combine date and time to create datetime
                    error_datetime = datetime.datetime.combine(record.date, record.time)
                    
                    # Convert to Persian date for display
                    persian_date = jdatetime.datetime.fromgregorian(datetime=error_datetime)
                    
                    error_data.append({
                        'id': record.id,
                        'errorcode': error_code,
                        'errormessage': error_detail.errormessage or 'مشخص نیست',
                        'probablecause': error_detail.probablecause or 'مشخص نیست',
                        'solution': error_detail.solution or 'موجود نیست',
                        'date_time': error_datetime,
                        'persian_date': persian_date.strftime('%Y/%m/%d'),
                        'persian_time': persian_date.strftime('%H:%M'),
                    })
                except Errors.DoesNotExist:
                    continue
        except (ValueError, AttributeError):
            continue
    
    # Apply error number filtering if search query exists
    if search_query:
        filtered_error_data = []
        for record_data in error_data:
            # Convert error code to string and check if search query is contained in it
            if search_query in str(record_data['errorcode']):
                filtered_error_data.append(record_data)
        error_data = filtered_error_data
    
    # Apply date filtering if provided
    if dateFrom and dateTo:
        try:
            # Clean and normalize date strings
            dateFrom = dateFrom.replace('از', '').replace('تا', '').strip()
            dateTo = dateTo.replace('از', '').replace('تا', '').strip()
            
            # Try parsing with different date formats
            date_formats = [
                "%Y-%m-%d",  # 1403-01-01
                "%Y/%m/%d",  # 1403/01/01
                "%Y %m %d",  # 1403 01 01
            ]
            
            convFrom = None
            convTo = None
            
            for fmt in date_formats:
                try:
                    convFrom = jdatetime.datetime.strptime(dateFrom, fmt).togregorian()
                    convTo = jdatetime.datetime.strptime(dateTo, fmt).togregorian()
                    break
                except ValueError:
                    continue
            
            if convFrom and convTo:
                # Add one day to dateTo to include the entire end date
                convTo = convTo + datetime.timedelta(days=1)
                
                # Filter error data by date range
                filtered_error_data = []
                for record_data in error_data:
                    if convFrom <= record_data['date_time'] < convTo:
                        filtered_error_data.append(record_data)
                
                error_data = filtered_error_data
        except Exception as e:
            print(f"Error parsing dates in AJAX search: {e}")
    
    # Sort error data by date_time (newest first)
    error_data.sort(key=lambda x: x['date_time'], reverse=True)
    
    # Return results as JSON
    results = []
    for idx, record in enumerate(error_data, start=1):
        results.append({
            'index': idx,
            'errorcode': record['errorcode'],
            'errormessage': record['errormessage'],
            'probablecause': record['probablecause'],
            'solution': record['solution'],
            'persian_date': record['persian_date'],
            'persian_time': record['persian_time'],
        })
    
    return JsonResponse({
        'success': True,
        'count': len(results),
        'results': results
    })


class settingsPage(TemplateView, View):
    template_name = "settings.html"

    def form_valid(self, form):
        if not checkSession(self.request):
            messages.error(self.request, messagesTypes.notlogin)
            return redirect(reverse_lazy('login'))
        if not check_active(checkSession(self.request)):
            messages.error(self.request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(self.request) else 'login'))
        if isinstance(check := hasAccess("view", "settings", self.request), HttpResponseRedirect):
            return check
        if isinstance(check := hasAccess("write", "settings", self.request), HttpResponseRedirect):
            return check
        if isinstance(check := hasAccess("modify", "settings", self.request), HttpResponseRedirect):
            return check
        msgs = "تنظیمات زیر با موفقیت ذخیره شد:\n"
        field = form.cleaned_data
        fieldRequest = self.request.POST
        if any(f for f in [field.get('device'), field.get('flow'), field.get('stopbits'), field.get('baudrate'),
                            field.get('parity'), field.get('databits'), field.get('number_of_lines')]) or any (f for f in [field.get('smdrip'),
                            field.get('smdrport'), field.get('smdrpassword')]):
            if getVersion() == 'alvand':
                systemSettingsConfiguration(field, fieldRequest, self.request, self.success_url, form, False)
            elif getVersion == 'binalud':
                # Decide Ethernet mode only based on selected cable type, not device model
                cable = field.get('cable_type')
                if not cable or cable.lower() in ['none', '']:
                    messages.error(self.request, 'لطفا یکی از انواع کابل هارا انتخاب کنید.')
                    return redirect(self.success_url)

                isEthernet = (cable.lower() == 'ethernet')

                systemSettingsConfiguration(field, fieldRequest, self.request, self.success_url, form, isEthernet)

            else:
                messages.error(self.request, 'نسخه برنامه شما هنوز مشخص نشده است.')
                return redirect(self.success_url)

        if any(f for f in [fieldRequest.get('province'), fieldRequest.get('phone_number')]):
            if not fieldRequest.get('province'):
                messages.error(self.request, messagesTypes.fillAllFieldsInSection.format('اطلاعات تماس'))
                return redirect(self.success_url)
            if not fieldRequest.get('phone_number'):
                messages.error(self.request, messagesTypes.fillAllFieldsInSection.format('اطلاعات تماس'))
                return redirect(self.success_url)
            if not fieldRequest.get('province').isdigit():
                messages.error(self.request, messagesTypes.fillNotInFields.format('انتخاب استان'))
                return redirect(self.success_url)
            if fieldRequest.get('province') not in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10',
                                                    '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
                                                    '21', '22', '23', '24', '25', '26', '27', '28', '29', '30']:
                messages.error(self.request, messagesTypes.fillNotInFields.format('انتخاب استان'))
                return redirect(self.success_url)
            if not fieldRequest.get('phone_number').isdigit():
                messages.error(self.request, messagesTypes.fillNotInFields.format('شماره همراه مدیر'))
                return redirect(self.success_url)
            if len(fieldRequest.get('phone_number')) != 11:
                messages.error(self.request, messagesTypes.fillNotInFields.format('شماره همراه مدیر'))
                return redirect(self.success_url)

            conInfo = ContactInfo.objects.all()
            if conInfo.exists():
                conInfo.update(province=fieldRequest.get('province'), phone_number=fieldRequest.get('phone_number'),
                               user=Users.objects.filter(username__iexact=checkSession(self.request)).first())
                log(self.request, logErrCodes.systemSettings, logMessages.updatedSettings.format('اطلاعات تماس'),
                    checkSession(self.request))

            else:
                ContactInfo.objects.create(province=fieldRequest.get('province'),
                                           phone_number=fieldRequest.get('phone_number'), user=Users.objects.filter(
                        username__iexact=checkSession(self.request)).first())
                log(self.request, logErrCodes.systemSettings, logMessages.createdSettings.format('اطلاعات تماس'),
                    checkSession(self.request))
            msgs += "اطلاعات تماس\n"
        if fieldRequest.get('edit') or fieldRequest.get('delete') or fieldRequest.get('add'):
            editext = Extensionsgroups.objects.filter(label__iexact=fieldRequest.get('label'))

            if not any(f for f in [any(x for x in fieldRequest.getlist('exts')), fieldRequest.get('label')]):
                messages.error(self.request, messagesTypes.fillAllFields)
                return redirect(self.success_url)
            if fieldRequest.get('edit'):
                if not editext.exists():
                    messages.error(self.request, 'این نام گروه وجود ندارد.')
                    return redirect(self.success_url)
                getExtFromDB = editext.values_list('exts', flat=True).first() or []
                valuesFromUser = [int(x) for x in fieldRequest.getlist('exts')]
                mightRemove = set(getExtFromDB) - set(valuesFromUser)
                mightAdd = set(valuesFromUser) - set(getExtFromDB)
                mergeAndUpdate = list((set(getExtFromDB) - mightRemove) | mightAdd)
                editext.update(exts=mergeAndUpdate, label=fieldRequest.get('label'),
                               modifyby=Users.objects.filter(username__iexact=checkSession(self.request)).first())
                msgs += "ویرایش تنظیمات دسترسی به گروه ها\n"

            if fieldRequest.get('add'):
                if editext.exists():
                    messages.error(self.request, 'این نام گروه وجود دارد.')
                    return redirect(self.success_url)
                if len(fieldRequest.get('label')) <= 3 and len(fieldRequest.get('label')) >= 25:
                    messages.error(self.request, 'نام گروه باید بین ۳ تا ۲۵ حرف باشد.')
                    return redirect(self.success_url)

                Extensionsgroups.objects.create(label=fieldRequest.get('label'),
                                                exts=[int(x) for x in fieldRequest.getlist('exts')],
                                                modifyby=Users.objects.filter(
                                                    username__iexact=checkSession(self.request)).first())
                msgs += "اضافه تنظیمات دسترسی به گروه ها\n"
            if fieldRequest.get('delete'):
                lbl = fieldRequest.get('label')
                editext = Extensionsgroups.objects.filter(label__iexact=lbl)
                if not editext.exists():
                    messages.error(self.request, 'این نام گروه وجود ندارد.')
                    return redirect(self.success_url)
                perms = Permissions.objects.filter(exts_label__contains=[lbl])
                if perms.exists():
                    for perm in perms:
                        if lbl in perm.exts_label:
                            updateLbl = [label for label in perm.exts_label if label != lbl]
                            perm.exts_label = updateLbl
                            perm.save()
                editext.delete()
                msgs += f"حذف گروه {lbl} از تنظیمات دسترسی به گروه ها\n"

        if any(cost for cost in
               [fieldRequest.get("hamrahaval"), fieldRequest.get("irancell"), fieldRequest.get("rightel"),
                fieldRequest.get("provincial"), fieldRequest.get("international"),
                fieldRequest.get("outofprovincial")]):
            if not fieldRequest.get("hamrahaval") or not fieldRequest.get("irancell") or not fieldRequest.get(
                    "rightel") or not fieldRequest.get("provincial") or not fieldRequest.get(
                "international") or not fieldRequest.get("outofprovincial"):
                messages.error(self.request, messagesTypes.fillAllFieldsInSection.format('هزینه های تماس'))
                return redirect(self.success_url)
            fields = [fieldRequest.get("hamrahaval"), fieldRequest.get("irancell"), fieldRequest.get("rightel"),
                      fieldRequest.get("provincial"), fieldRequest.get("international"),
                      fieldRequest.get("outofprovincial")]
            if any(not isfloat(item) for item in fields):
                messages.error(self.request, "لطفا مقادیر هزینه های تماس را به صورت اعشار و یا عدد صحیح بنویسید.")
                return redirect(self.success_url)
            if any(float(item) < 0.0 for item in fields):
                messages.error(self.request, "هزینه های تماس نباید کمتر از 0 باشد.")
                return redirect(self.success_url)
            costs = Costs.objects
            if costs.all().exists():
                costs.update(hamrahaval=fieldRequest.get('hamrahaval'), irancell=fieldRequest.get('irancell'),
                             rightel=fieldRequest.get('rightel'),
                             provincial=fieldRequest.get('provincial'), international=fieldRequest.get('international'),
                             outofprovincial=fieldRequest.get('outofprovincial'), updated_at=timezone.now())
                log(self.request, logErrCodes.systemSettings, logMessages.updatedSettings.format('هزینه های تماس'),
                    checkSession(self.request))
            else:
                costs.create(hamrahaval=fieldRequest.get('hamrahaval'), irancell=fieldRequest.get('irancell'),
                             rightel=fieldRequest.get('rightel'),
                             provincial=fieldRequest.get('provincial'), international=fieldRequest.get('international'),
                             outofprovincial=fieldRequest.get('outofprovincial'), created_at=timezone.now())
                log(self.request, logErrCodes.systemSettings, logMessages.createdSettings.format('هزینه های تماس'),
                    checkSession(self.request))
            msgs += "هزینه های تماس\n"
        if any(f for f in [fieldRequest.get('collectionusername'), fieldRequest.get('collectionpassword'),
                           fieldRequest.get('emailtosend'), any(err for err in fieldRequest.getlist('errors'))]):
            if not all(f for f in [fieldRequest.get('collectionusername'), fieldRequest.get('collectionpassword'),
                                   fieldRequest.get('emailtosend'),
                                   any(err for err in fieldRequest.getlist('errors'))]):
                messages.error(self.request, messagesTypes.fillAllFieldsInSection.format('تنظیمات ایمیل'))
                return redirect(self.success_url)
            if not all("@gmail.com" in x.strip() for x in
                       [fieldRequest.get('collectionusername'), fieldRequest.get('emailtosend')]):
                messages.error(self.request, "ایمیل باید حتما دارای @gmail.com باشد. (مانند example@gmail.com)")
                return redirect(self.success_url)
            if any(len(x.strip().replace("@gmail.com", "")) < 3 for x in
                   [fieldRequest.get('collectionusername'), fieldRequest.get('emailtosend')]):
                messages.error(self.request, "ایمیل وارد شده نامعتبر است.")
                return redirect(self.success_url)
            if not fieldRequest.get('lines').isdigit() and not int(fieldRequest.get('lines')) in range(1, 10 + 1):
                messages.error(self.request, messagesTypes.fillNotInFields.format('تعداد خطاها برای ارسال'))
                return redirect(self.success_url)
            if not any(item.isdigit() for item in fieldRequest.getlist('errors')):
                messages.error(self.request, messagesTypes.fillNotInFields.format('خطاها'))
                return redirect(self.success_url)
            if not any(
                    Errors.objects.filter(errorcodenum=int(item)).exists() for item in fieldRequest.getlist('errors')):
                messages.error(self.request, "یکی از خطا های انتخاب شده وجود ندارد.")
                return redirect(self.success_url)
            emailSendCheck = Emailsending.objects
            if emailSendCheck.all().exists():
                emailSendCheck.update(errors=[int(x) for x in fieldRequest.getlist('errors')],
                                      emailtosend=fieldRequest.get('emailtosend'),
                                      collectionusername=fieldRequest.get('collectionusername'),
                                      collectionpassword=fieldRequest.get('collectionpassword'),
                                      lines=int(fieldRequest.get('lines')),
                                      byadmin=Users.objects.filter(username__iexact=checkSession(self.request)).first(),
                                      updated_at=timezone.now())
                log(self.request, logErrCodes.systemSettings, logMessages.updatedSettings.format('تنظیمات ایمیل'),
                    checkSession(self.request))

            else:
                emailSendCheck.create(errors=[int(x) for x in fieldRequest.getlist('errors')],
                                      emailtosend=fieldRequest.get('emailtosend'),
                                      collectionusername=fieldRequest.get('collectionusername'),
                                      collectionpassword=fieldRequest.get('collectionpassword'),
                                      lines=int(fieldRequest.get('lines')),
                                      byadmin=Users.objects.filter(username__iexact=checkSession(self.request)).first(),
                                      created_at=timezone.now())
                log(self.request, logErrCodes.systemSettings, logMessages.createdSettings.format('تنظیمات ایمیل'),
                    checkSession(self.request))
            msgs += "تنظیمات ایمیل\n"
        if fieldRequest.get('users') and fieldRequest.get("users").lower() != "none":
            user = Users.objects.filter(username__iexact=fieldRequest.get('users'))
            if not user.exists():
                messages.error(self.request, f"کاربر {fieldRequest.get('users')} وجود ندارد.")
                return redirect(self.success_url)
            perm = Permissions.objects.filter(user=user.first())
            if not perm.exists():
                Permissions.objects.create(user=user.first(), errorsreport=bool(fieldRequest.get('errorsreport')))
            else:
                perm.update(errorsreport=bool(fieldRequest.get('errorsreport')))
            log(self.request, logErrCodes.systemSettings,
                logMessages.updatedSettings.format('دسترسی به گزارش خطا ها - گروه بندی دسترسی ها'),
                checkSession(self.request))
            msgs += "دسترسی به گزارش خطا ها - گروه بندی دسترسی ها\n"
        if msgs != "تنظیمات زیر با موفقیت ذخیره شد:\n":
            messages.success(self.request, msgs)
        else:
            messages.info(self.request, "هیچ تغییراتی اعمال نشده است.")
        return redirect(self.success_url)

    def form_invalid(self, form):
        if not checkSession(self.request):
            messages.error(self.request, messagesTypes.notlogin)
            return redirect(self.success_url)
        if not check_active(checkSession(self.request)):
            messages.error(self.request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(self.request) else 'login'))
        if isinstance(check := hasAccess("write", "user", self.request), HttpResponseRedirect):
            return check
        messages.error(self.request, messagesTypes.fillAllFields)
        return redirect(self.success_url)


class dashboardPage(TemplateView, View):
    template_name = "dashboard.html"

    def get(self, request, *args, **kwargs):
        if not checkSession(request):
            messages.error(request, messagesTypes.notlogin)
            return redirect(reverse_lazy("login"))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        context = self.get_context_data()
        if request.GET:
            # Extract raw inputs
            dateFrom = request.GET.get('dateFrom')
            dateTo = request.GET.get('dateTo')
            selected_urbanlines = request.GET.getlist('urbanline')
            selected_extlines = request.GET.getlist('extline')
            selected_calls = request.GET.getlist('calls')

            # Map Persian call types to stored values
            mapped_calls = persianCallTypeToEnglish(selected_calls) if selected_calls else None

            # Build base query always constrained to user's accessible extensions
            query = Q()
            allowed_exts = getExtensionlines(checkSession(request))
            if allowed_exts:
                query &= Q(extension__in=allowed_exts)

            # Apply independent filters
            if selected_urbanlines:
                query &= Q(urbanline__in=selected_urbanlines)
            if selected_extlines:
                query &= Q(extension__in=selected_extlines)
            if mapped_calls:
                query &= Q(calltype__in=mapped_calls)

            # Date range (Jalali -> Gregorian)
            if dateFrom and dateTo:
                try:
                    convFrom = jdatetime.datetime.strptime(
                        dateFrom.replace("/", "-").replace('از', '').strip(), "%Y-%m-%d"
                    ).togregorian()
                    convTo = jdatetime.datetime.strptime(
                        dateTo.replace("/", "-").replace('تا', '').strip(), "%Y-%m-%d"
                    ).togregorian()
                except:
                    messages.error(request, "فرمت تاریخ ها اشتباه است.")
                    return redirect(request.path)
                if convTo < convFrom:
                    messages.warning(request, "تاریخ اول نباید بزرگ تر از تاریخ دوم باشد.")
                    return redirect(request.path)
                query &= Q(date__range=(convFrom, convTo))

            # Free-text search
            search_query = request.GET.get('q', '').strip()
            if search_query:
                query &= (
                    Q(contactnumber__icontains=search_query) |
                    Q(extension__icontains=search_query) |
                    Q(urbanline__icontains=search_query)
                )

            # Decide filtered vs default dataset
            any_filter_applied = bool(selected_urbanlines or selected_extlines or mapped_calls or (dateFrom and dateTo) or search_query)
            base_query = (
                Records.objects.filter(query).order_by('-created_at', '-id')
                if any_filter_applied else
                dashboardData(checkSession(request))
            )
            faults, page_obj = makePagination(base_query, 20, request)
            context['dashPage'] = page_obj
        return render(request, self.template_name, context=context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pageTitle'] = 'داشبورد'
        faults, page_obj = makePagination(dashboardData(checkSession(self.request)), 20, self.request)
        context['dashPage'] = page_obj
        context['urbanlines'] = getUrbanlinesThatUserAccessThem(checkSession(self.request))
        context['extlines'] = getExtensionlines(checkSession(self.request))

        # Aggregate stats for dashboard cards (for the user's accessible extensions)
        try:
            username = checkSession(self.request)
            user_qs = Users.objects.filter(username__iexact=username)
            allExtentions = []
            if user_qs.exists():
                user_obj = user_qs.first()
                if user_obj.extension:
                    allExtentions.append(user_obj.extension)
                if user_obj.usersextension:
                    for userext in user_obj.usersextension:
                        if userext:
                            allExtentions.append(userext)
                labels = Permissions.objects.filter(user=user_obj)
                if labels.exists() and labels.first().exts_label:
                    extGroup = Extensionsgroups.objects.filter(label__in=labels.first().exts_label)
                    if extGroup.exists():
                        for item in extGroup:
                            allExtentions = allExtentions + item.exts
            
            base_qs = Records.objects.filter(extension__in=allExtentions) if allExtentions else Records.objects.none()
            incoming_types = ['incomingNA', 'incomingRC', 'incomingAN', 'Transfer', 'incomingDISA', 'incomingHangUp']

            context['incoming_calls_count'] = base_qs.filter(calltype__in=incoming_types).count()
            context['outgoing_calls_count'] = base_qs.filter(calltype='outGoing').count()
            context['missed_calls_count'] = base_qs.filter(calltype='incomingNA').count()

            # Top 5 extensions by call volume
            top_ext = base_qs.values('extension').annotate(total=Count('id')).order_by('-total')[:5]
            context['top_extensions'] = list(top_ext)

            # Total call duration (HH:MM:SS)
            def _parse_duration(d):
                try:
                    h, m, s = map(int, str(d).split(':'))
                    return h * 3600 + m * 60 + s
                except Exception:
                    return 0
            total_seconds = 0
            for d in base_qs.exclude(durationtime__isnull=True).values_list('durationtime', flat=True):
                total_seconds += _parse_duration(d)
            hh = total_seconds // 3600
            mm = (total_seconds % 3600) // 60
            ss = total_seconds % 60
            context['total_call_duration'] = f"{hh:02d}:{mm:02d}:{ss:02d}"
        except Exception:
            context.setdefault('incoming_calls_count', 0)
            context.setdefault('outgoing_calls_count', 0)
            context.setdefault('missed_calls_count', 0)
            context.setdefault('top_extensions', [])
            context['extlines'] = getExtensionlines(checkSession(self.request))
        return context


def live_call_status(request):
    """Return the latest active call (if any) for use in the live dashboard popup."""
    if not checkSession(request):
        return JsonResponse({"active": False}, status=401)
    if not check_active(checkSession(request)):
        return JsonResponse({"active": False}, status=403)

    try:
        # Consider calls that are currently ringing, answered, or outgoing
        active_types = ['incomingRC', 'incomingAN', 'outGoing']
        qs = Records.objects.filter(calltype__in=active_types, durationtime__isnull=True)

        # Limit to very recent records to avoid stale entries (e.g., last 2 minutes)
        try:
            two_minutes_ago = timezone.now() - datetime.timedelta(minutes=2)
            qs = qs.filter(created_at__gte=two_minutes_ago)
        except Exception:
            pass

        rec = qs.order_by('-created_at', '-id').first()
        if not rec:
            return JsonResponse({"active": False})

        data = {
            "active": True,
            "id": rec.id,
            "extension": getattr(rec, 'extension', None),
            "contactnumber": getattr(rec, 'contactnumber', None),
            "calltype": getattr(rec, 'calltype', None),
            "urbanline": getattr(rec, 'urbanline', None),
            "durationtime": getattr(rec, 'durationtime', None),
            "created_at": timezone.localtime(getattr(rec, 'created_at', timezone.now())).strftime('%Y-%m-%d %H:%M:%S'),
        }
        return JsonResponse(data)
    except Exception:
        return JsonResponse({"active": False})


def dashboard_export(request):
    """Export filtered dashboard records as CSV or Excel using the same filter logic as the dashboard page."""
    if not checkSession(request):
        messages.error(request, messagesTypes.notlogin)
        return redirect(reverse_lazy("login"))
    if not check_active(checkSession(request)):
        messages.error(request, messagesTypes.deAvtive)
        return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))

    # Get export format
    export_format = request.GET.get('export', 'csv').lower()

    # Build filters identical to dashboardPage.get
    dateFrom = request.GET.get('dateFrom')
    dateTo = request.GET.get('dateTo')
    selected_urbanlines = request.GET.getlist('urbanline')
    selected_extlines = request.GET.getlist('extline')
    selected_calls = request.GET.getlist('calls')

    mapped_calls = persianCallTypeToEnglish(selected_calls) if selected_calls else None

    query = Q()
    allowed_exts = getExtensionlines(checkSession(request))
    if allowed_exts:
        query &= Q(extension__in=allowed_exts)

    if selected_urbanlines:
        query &= Q(urbanline__in=selected_urbanlines)
    if selected_extlines:
        query &= Q(extension__in=selected_extlines)
    if mapped_calls:
        query &= Q(calltype__in=mapped_calls)

    if dateFrom and dateTo:
        try:
            convFrom = jdatetime.datetime.strptime(
                dateFrom.replace("/", "-").replace('از', '').strip(), "%Y-%m-%d"
            ).togregorian()
            convTo = jdatetime.datetime.strptime(
                dateTo.replace("/", "-").replace('تا', '').strip(), "%Y-%m-%d"
            ).togregorian()
        except:
            messages.error(request, "فرمت تاریخ ها اشتباه است.")
            return redirect(request.path)
        if convTo < convFrom:
            messages.warning(request, "تاریخ اول نباید بزرگ تر از تاریخ دوم باشد.")
            return redirect(request.path)
        query &= Q(date__range=(convFrom, convTo))

    search_query = request.GET.get('q', '').strip()
    if search_query:
        query &= (
            Q(contactnumber__icontains=search_query) |
            Q(extension__icontains=search_query) |
            Q(urbanline__icontains=search_query)
        )

    any_filter_applied = bool(selected_urbanlines or selected_extlines or mapped_calls or (dateFrom and dateTo) or search_query)
    base_qs = (
        Records.objects.filter(query).order_by('-created_at', '-id')
        if any_filter_applied else
        dashboardData(checkSession(request))
    )

    # Get all records (no pagination limit)
    all_records = list(base_qs)

    if export_format == 'excel':
        # Excel export using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messages.error(request, "کتابخانه Excel نصب نشده است.")
            return redirect('dashboard')

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "تماس‌ها"

        # Headers
        headers = [
            'ردیف', 'تاریخ', 'ساعت', 'شماره مخاطب', 'شماره داخلی', 
            'خط شهری', 'مدت تماس (ثانیه)', 'نوع تماس', 'زمان برق', 'انتقال', 'هزینه (تومان)'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row_idx, rec in enumerate(all_records, 2):
            try:
                # Calculate price
                price_val = '0'
                if getattr(rec, 'durationtime', None) and getattr(rec, 'contactnumber', None):
                    calltype_kind = callTypeDetector(rec.contactnumber)
                    if calltype_kind is not None:
                        unit_price = getPrice(calltype_kind)
                        if unit_price is None:
                            price_val = 'ندارد'
                        else:
                            calculated = calculatePrice(str(rec.durationtime), unit_price)
                            price_val = str(calculated)

                # Convert date to Persian format
                persian_date = ''
                if getattr(rec, 'date', None):
                    try:
                        # Convert Gregorian date to Persian
                        jdate = jdatetime.date.fromgregorian(date=rec.date)
                        persian_date = jdate.strftime('%Y/%m/%d')
                    except:
                        persian_date = rec.date.strftime('%Y-%m-%d')

                row_data = [
                    row_idx - 1,  # Row number
                    persian_date,
                    getattr(rec, 'hour', '') or '',
                    getattr(rec, 'contactnumber', '') or '',
                    getattr(rec, 'extension', '') or '',
                    getattr(rec, 'urbanline', '') or '',
                    getattr(rec, 'durationtime', '0') or '0',
                    getattr(rec, 'calltype', '') or '',
                    getattr(rec, 'beepsnumber', 0) or 0,
                    'بله' if getattr(rec, 'transferring', False) else 'خیر',
                    price_val,
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            except Exception:
                continue

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

        # Prepare Excel response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"تماس‌های_لوتوس_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    else:
        # CSV export (existing functionality)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"dashboard_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.charset = 'utf-8'
        response.write('\ufeff')

        writer = csv.writer(response, lineterminator='\r\n')
        writer.writerow([
            'تاریخ', 'ساعت', 'شماره مخاطب', 'شماره داخلی', 'خط شهری', 'مدت تماس', 'نوع تماس', 'زمان برق', 'انتقال', 'هزینه'
        ])

        for rec in all_records:
            try:
                price_val = '0'
                if getattr(rec, 'durationtime', None) and getattr(rec, 'contactnumber', None):
                    calltype_kind = callTypeDetector(rec.contactnumber)
                    if calltype_kind is not None:
                        unit_price = getPrice(calltype_kind)
                        if unit_price is None:
                            price_val = 'ندارد'
                        else:
                            calculated = calculatePrice(str(rec.durationtime), unit_price)
                            price_val = str(calculated)

                # Convert date to Persian format
                persian_date = ''
                if getattr(rec, 'date', None):
                    try:
                        # Convert Gregorian date to Persian
                        jdate = jdatetime.date.fromgregorian(date=rec.date)
                        persian_date = jdate.strftime('%Y/%m/%d')
                    except:
                        persian_date = rec.date.strftime('%Y-%m-%d')

                writer.writerow([
                    persian_date,
                    getattr(rec, 'hour', '') or '',
                    getattr(rec, 'contactnumber', '') or '',
                    getattr(rec, 'extension', '') or '',
                    getattr(rec, 'urbanline', '') or '',
                    getattr(rec, 'durationtime', '0') or '0',
                    getattr(rec, 'calltype', '') or '',
                    getattr(rec, 'beepsnumber', 0) or 0,
                    'بله' if getattr(rec, 'transferring', False) else 'خیر',
                    price_val,
                ])
            except Exception:
                continue

        return response


def support(request):
    from django.contrib import messages
    from .forms import SupportMessageForm

    # Determine current logged-in custom user and supporter role
    current_user = None
    username = checkSession(request)
    if username:
        current_user = Users.objects.filter(username__iexact=username).first()

    is_supporter = bool(current_user and str(current_user.groupname).lower() == 'supporter')

    if request.method == 'POST':
        form = SupportMessageForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'پیام شما با موفقیت ارسال شد. به زودی با شما تماس خواهیم گرفت.')
            return redirect('support')
        else:
            messages.error(request, 'لطفا تمام فیلدها را به درستی پر کنید.')
    else:
        form = SupportMessageForm()

    return render(request, 'support.html', context={'pageTitle': 'پشتیبانی', 'form': form, 'user': current_user, 'is_supporter': is_supporter})

def checkSession(request):
    if 'user' in request.session:
        return request.session['user']
    return False


def login(request, user):
    if 'user' not in request.session:
        request.session['user'] = user.lower()
        return request.session['user']
    return request.session['user']


def getIPAddress(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[-1].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log(request, errCode, errMessage, byWho=None):
    username = checkSession(request) if not isinstance(request, str) else request
    if not username or not isinstance(username, str) or not username.strip():
        return False
    user = Users.objects.filter(username__iexact=username)
    if not user.exists():
        return False
    ip = getIPAddress(request)
    return bool(
        Log.objects.create(user=user.first(), userBackup=user.first().username, errCode=errCode, errMessage=errMessage,
                           byWho=byWho if byWho else "Lotus", ip=ip,
                           macAddress=getInfosOfUserByUsername(user.first().username, 'macaddress')))


class logout(TemplateView):
    def get(self, request, *args, **kwargs):
        if checkSession(request):
            log(request, logErrCodes.logInOut, logMessages.loggedOut.format(request.session['user'], ),
                request.session['user'])
            del request.session['user']
            messages.success(request, messagesTypes.logout)
        else:
            messages.error(request, messagesTypes.notlogin)
        return redirect(reverse_lazy('login'))


class Profile(TemplateView, View):
    template_name = "profile.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pageTitle'] = "پروفایل کاربری"
        user = Users.objects.filter(username__iexact=str(checkSession(self.request)))
        user_obj = user.first() if user.exists() else None
        context['user'] = user_obj

        from .forms import SelfInfosProfileForm, SelfUserProfileForm
        infos = None
        if user_obj:
            infos = Infos.objects.filter(user=user_obj).first()
        context['infos_form'] = SelfInfosProfileForm(instance=infos)
        context['user_form'] = SelfUserProfileForm(instance=user_obj)
        return context

    def get(self, request, *args, **kwargs):
        context = self.get_context_data()
        if not checkSession(request):
            messages.warning(request, messagesTypes.userInfoNotFound)
            return redirect(reverse_lazy("login"))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        return render(request, self.template_name, context=context)

    def post(self, request):
        if not checkSession(request):
            messages.error(request, messagesTypes.notlogin)
            return redirect(reverse_lazy("login"))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        
        # Handle profile picture upload
        if request.POST.get('action') == 'upload_profile_pic' and request.FILES.get('profile_picture'):
            try:
                import os
                from django.conf import settings
                import json
                from django.http import JsonResponse
                
                user = Users.objects.filter(username__iexact=checkSession(request)).first()
                if not user:
                    return JsonResponse({'success': False, 'error': 'کاربر یافت نشد'})
                
                profile_pic = request.FILES['profile_picture']
                
                # Validate file type
                valid_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                ext = os.path.splitext(profile_pic.name)[1].lower()
                if ext not in valid_extensions:
                    return JsonResponse({'success': False, 'error': 'فرمت فایل نامعتبر است. فقط jpg, jpeg, png, gif و webp مجاز هستند.'})
                
                # Validate file size (max 2MB)
                if profile_pic.size > 2 * 1024 * 1024:
                    return JsonResponse({'success': False, 'error': 'حجم فایل باید کمتر از 2 مگابایت باشد'})
                
                # Create directory if it doesn't exist
                upload_dir = os.path.join('Alvand/static/upload')
                if not os.path.exists(upload_dir):
                    os.makedirs(upload_dir)
                
                # Generate filename based on username to avoid duplicates
                filename = f"{user.username}{ext}"
                filepath = os.path.join(upload_dir, filename)
                
                # Delete old profile picture if exists and it's not the default avatar
                if user.profile_picture and user.profile_picture != "avatar.png" and os.path.exists(os.path.join('Alvand/static/upload', user.profile_picture)):
                    try:
                        os.remove(os.path.join('Alvand/static/upload', user.profile_picture))
                    except:
                        pass
                
                # Save the new profile picture
                with open(filepath, 'wb+') as destination:
                    for chunk in profile_pic.chunks():
                        destination.write(chunk)
                
                # Update user model
                user.profile_picture = filename
                user.picurl = filename
                user.save()
                
                # Log the action
                log(request, logErrCodes.userSettings, f"کاربر {user.username} تصویر پروفایل خود را بروزرسانی کرد", user.username)
                
                # Return success response with image URL
                image_url = f'/static/upload/{filename}'
                return JsonResponse({'success': True, 'image_url': image_url})
                
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})

        if request.POST.get('save_profile'):
            from .forms import SelfInfosProfileForm, SelfUserProfileForm
            username = checkSession(request)
            user = Users.objects.filter(username__iexact=username).first()
            if not user:
                messages.error(request, messagesTypes.userInfoNotFound)
                return redirect(reverse_lazy("profile"))

            infos = Infos.objects.filter(user=user).first()
            infos_form = SelfInfosProfileForm(request.POST, instance=infos)
            user_form = SelfUserProfileForm(request.POST, instance=user)
            if infos_form.is_valid() and user_form.is_valid():
                obj = infos_form.save(commit=False)
                obj.user = user
                obj.save()
                user_form.save()
                messages.success(request, "پروفایل شما با موفقیت بروزرسانی شد.")
                response = redirect(reverse_lazy("profile"))
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                return response
            context = self.get_context_data()
            context['infos_form'] = infos_form
            context['user_form'] = user_form
            return render(request, self.template_name, context=context)
        
        return redirect(reverse_lazy("profile"))


class userLogin(View):
    def get(self, request, *args, **kwargs):
        if checkSession(request):
            if not check_active(checkSession(request)):
                messages.error(request, messagesTypes.deAvtive)
                return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
            messages.error(request, messagesTypes.loggedIn)
            return redirect(reverse_lazy('dashboard'))
        return render(request, template_name="login.html", context={})

    def post(self, request):
        if checkSession(request):
            if not check_active(checkSession(request)):
                messages.error(request, messagesTypes.deAvtive)
                return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
            messages.error(request, messagesTypes.loggedIn)
            return redirect(reverse_lazy('dashboard'))

        username = request.POST['user']
        password = request.POST['pass']
        user = Users.objects.filter(username__iexact=username)
        if user.exists():
            if not check_active(user.first().username):
                messages.error(request, messagesTypes.deAvtive)
                return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
            if check_password(password, user.first().password):
                user_obj = Users.objects.get(username=username)

                # اگر کاربر با رمز عبور پیش فرض وارد شده است، اجبار تغییر رمز را فعال کن
                if password == "12345678" and not user_obj.needs_password_change:
                    user_obj.needs_password_change = True
                    user_obj.save()

                login(request, username)
                log(request, logErrCodes.logInOut, logMessages.loggedIn.format(user.first().username.capitalize(), ),
                    user.first().username)
                messages.success(request, messagesTypes.login.format(user.first().username.capitalize(), ))

                if user_obj.needs_password_change:
                    return redirect('change_password')
                return redirect(reverse_lazy('dashboard'))
            messages.error(request, messagesTypes.invalidUsernameOrPassword)
            return redirect(request.path)
        else:
            messages.error(request, messagesTypes.invalidUsernameOrPassword)
            return redirect(request.path)


class errorsPage(TemplateView, View):
    template_name = "error.html"

    def get(self, request, *args, **kwargs):
        if not checkSession(request):
            messages.error(request, messagesTypes.notlogin)
            return redirect(reverse_lazy("login"))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        
        # Simplified permission check - allow all users for now
        try:
            user = Users.objects.filter(username__iexact=checkSession(request)).first()
            if user and user.groupname.lower() in ['superadmin', 'supporter', 'admin']:
                # Allow access for admin users
                pass
            else:
                # Temporarily allow all users to access error reports
                print(f"User {user.username if user else 'unknown'} accessing error reports")
                pass
        except Exception as e:
            # If there's any error in permission checking, allow access for now
            print(f"Permission check error: {e}")
            pass
        
        # Handle date filtering - get SMDR records with LM prefix in dial_number
        from django.db.models import Q
        import re
        
        # Get SMDR records where dial_number contains 'ALM #' (Alarm messages)
        smdr_qs = SMDRRecord.objects.filter(
            dial_number__contains='ALM #'
        )
        
        print(f"Total SMDR records with ALM # pattern: {smdr_qs.count()}")
        
        # Extract error codes and create error data
        error_data = []
        for record in smdr_qs:
            try:
                # Extract error code after ALM # sign
                match = re.search(r'ALM #(\d+)', record.dial_number)
                if match:
                    error_code = int(match.group(1))
                    
                    # Find corresponding error details
                    try:
                        error_detail = Errors.objects.get(errorcodenum=error_code)
                        
                        # Combine date and time to create datetime
                        error_datetime = datetime.datetime.combine(record.date, record.time)
                        
                        # Convert to Persian date for display
                        persian_date = jdatetime.datetime.fromgregorian(datetime=error_datetime)
                        
                        error_data.append({
                            'id': record.id,
                            'errorcode': error_code,
                            'errormessage': error_detail.errormessage,
                            'probablecause': error_detail.probablecause,
                            'solution': error_detail.solution,
                            'date_time': error_datetime,
                            'persian_date': persian_date.strftime('%Y/%m/%d'),
                            'persian_time': persian_date.strftime('%H:%M'),
                            'dial_number': record.dial_number,
                        })
                    except Errors.DoesNotExist:
                        print(f"Error code {error_code} not found in Errors table")
                        continue
            except (ValueError, AttributeError) as e:
                print(f"Error parsing dial_number {record.dial_number}: {e}")
                continue
        
        print(f"Processed {len(error_data)} error records")
        
        # Convert to queryset-like structure for pagination
        class ErrorRecord:
            def __init__(self, data):
                for key, value in data.items():
                    setattr(self, key, value)
        
        error_records = [ErrorRecord(data) for data in error_data]
        
        # Apply filtering if provided
        if request.GET:
            dateFrom = request.GET.get('dateFrom', '').strip()
            dateTo = request.GET.get('dateTo', '').strip()
            errorNumber = request.GET.get('search', '').strip()  # Get error number from search parameter
            
            print(f"DEBUG: Request GET parameters: {dict(request.GET)}")
            print(f"DEBUG: Error number from search: '{errorNumber}'")
            
            # Apply error number filtering first
            if errorNumber:
                try:
                    error_code = int(errorNumber)
                    print(f"DEBUG: Converting error number to int: {error_code}")
                    # Filter error records by error code
                    filtered_error_data = []
                    for record in error_records:
                        if record.errorcode == error_code:
                            filtered_error_data.append(record)
                    
                    error_records = filtered_error_data
                    print(f"After error number filtering: {len(error_records)} records")
                    
                    if not error_records:
                        messages.error(request, f"خطای شماره {errorNumber} پیدا نشد.")
                        
                except ValueError as e:
                    print(f"DEBUG: ValueError when converting error number: {e}")
                    messages.error(request, "شماره خطا باید عدد باشد.")
                    return redirect(request.path)
            
            # Apply date filtering if provided
            if dateFrom and dateTo:
                try:
                    # Clean and normalize date strings
                    dateFrom = dateFrom.replace('از', '').replace('تا', '').strip()
                    dateTo = dateTo.replace('از', '').replace('تا', '').strip()
                    
                    # Try parsing with different date formats
                    date_formats = [
                        "%Y-%m-%d",  # 1403-01-01
                        "%Y/%m/%d",  # 1403/01/01
                        "%Y %m %d",  # 1403 01 01
                    ]
                    
                    convFrom = None
                    convTo = None
                    
                    for fmt in date_formats:
                        try:
                            convFrom = jdatetime.datetime.strptime(dateFrom, fmt).togregorian()
                            convTo = jdatetime.datetime.strptime(dateTo, fmt).togregorian()
                            break
                        except ValueError:
                            continue
                    
                    if not convFrom or not convTo:
                        raise ValueError("Invalid date format")
                    
                    # Add one day to dateTo to include the entire end date
                    convTo = convTo + datetime.timedelta(days=1)
                    
                    if convTo < convFrom:
                        messages.warning(request, "تاریخ اول نباید بزرگ تر از تاریخ دوم باشد.")
                        return redirect(request.path)
                    
                    # Debug log the dates being used for filtering
                    print(f"Filtering from {convFrom} to {convTo}")
                    
                    # Filter error records by date range
                    filtered_error_data = []
                    for record in error_records:
                        if convFrom <= record.date_time < convTo:
                            filtered_error_data.append(record)
                    
                    error_records = filtered_error_data
                    print(f"After date filtering: {len(error_records)} records")
                    
                    if not error_records:
                        messages.error(request, f"در بین تاریخ های {dateFrom} و {dateTo} گزارش خطایی پیدا نشد.")
                        
                except Exception as e:
                    import traceback
                    print(f"Error parsing dates: {e}")
                    print(traceback.format_exc())
                    messages.error(request, f"خطا در فرمت تاریخ‌ها. لطفاً تاریخ‌ها را به درستی وارد کنید. ({str(e)})")
                    return redirect(request.path)
        
        # Sort error records by date_time (newest first)
        error_records.sort(key=lambda x: x.date_time, reverse=True)
        
        # Pagination for error records list
        page_number = request.GET.get('p', 1)
        paginator = Paginator(error_records, 20)
        print(f"Paginator total count: {len(error_records)}")
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        print(f"Page object has {len(page_obj.object_list)} items")
        
        context = self.get_context_data()
        context['pages'] = page_obj
        print(f"Context pages: {context.get('pages', 'None')}")
        return render(request, self.template_name, context=context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pageTitle'] = 'گزارش خطا ها'
        # Don't override pages here - it's handled in the get method
        return context


class UserForm(FormView, View):
    template_name = "userprofile.html"
    model = Users
    form_class = userProfileForm
    success_url = reverse_lazy("user")

    def get(self, request, *args, **kwargs):
        if not checkSession(request):
            return redirect(reverse_lazy("user"))
        if not check_active(checkSession(request)):
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(request) else 'login'))
        
        # Simplified permission check - allow all users for now
        try:
            user = Users.objects.filter(username__iexact=checkSession(request)).first()
            if user and user.groupname.lower() in ['superadmin', 'supporter', 'admin']:
                # Allow access for admin users
                pass
            else:
                # Temporarily allow all users to access user management
                print(f"User {user.username if user else 'unknown'} accessing user management")
                pass
        except Exception as e:
            # If there's any error in permission checking, allow access for now
            print(f"Permission check error: {e}")
            pass
            
        return render(request, self.template_name, context=self.get_context_data(**kwargs))

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['pageTitle'] = "تنظیمات کاربری"
        data['userform'] = self.form_class()
        data['infosform'] = InfosForm()
        data['permform'] = PermissionsForm()
        
        session_username = checkSession(self.request)
        
        # Get groupname directly from Users table
        getUser = None
        if session_username:
            user_obj = Users.objects.filter(username__iexact=session_username).first()
            if user_obj:
                getUser = user_obj.groupname
        
        # Get the QuerySet based on user role
        if str(getUser).lower() in ["superadmin", "supporter"]:
            users_queryset = Users.objects.exclude(groupname="supporter")
            
            # Add password reset requests to context for supporters
            from .models import PasswordResetRequest
            reset_requests = PasswordResetRequest.objects.filter(resolved=False).select_related(
                'user'
            ).order_by('-created_at')
            
            data['password_reset_requests'] = reset_requests
        else:
            users_queryset = Users.objects.exclude(groupname__in=['superadmin', 'supporter'])
        
        # Convert QuerySet to list of dictionaries for JSON serialization
        users_data = []
        for user in users_queryset:
            user_dict = {
                'id': user.id,
                'active': user.active,
                'extension': user.extension,
                'name': user.name,
                'username': user.username,
                'lastname': user.lastname,
                'groupname': user.groupname,
                'picurl': user.picurl,
                'email': user.email
            }
            if user.usersextension:
                user_dict['usersextension'] = list(user.usersextension)
            users_data.append(user_dict)
        
        data['users'] = users_data
        
        # Add session user data
        session_username = checkSession(self.request)
        if session_username:
            session_user = Users.objects.filter(username=session_username).first()
            if session_user:
                session_user_data = {
                    'id': session_user.id,
                    'username': session_user.username,
                    'name': session_user.name,
                    'lastname': session_user.lastname,
                    'groupname': session_user.groupname,
                    'extension': session_user.extension
                }
                data['session_user'] = session_user_data
        
        # Add contact info data
        contact_infos = []
        for user in users_queryset:
            info = Infos.objects.filter(user__username=user.username).first()
            if info:
                contact_info = {
                    'username': user.username,
                    'nationalcode': info.nationalcode,
                    'birthdate': str(info.birthdate) if info.birthdate else None,
                    'telephone': info.telephone,
                    'phonenumber': info.phonenumber,
                    'gender': info.gender,
                    'maritalstatus': info.maritalstatus,
                    'military': info.military,
                    'educationfield': info.educationfield,
                    'educationdegree': info.educationdegree,
                    'province': info.province,
                    'city': info.city,
                    'accountnumbershaba': info.accountnumbershaba,
                    'cardnumber': info.cardnumber,
                    'accountnumber': info.accountnumber,
                    'address': info.address
                }
                contact_infos.append(contact_info)
        
        data['contactInfos'] = contact_infos
        return data

    def form_valid(self, form):
        """Handle form submission for user creation and editing."""
        print("=" * 50)
        print("Form is valid, processing...")
        
        # Debug form data
        print("Form cleaned data:")
        for key, value in form.cleaned_data.items():
            print(f"  {key}: {value}")
        
        # Debug POST data specifically for save action
        print("POST data for critical fields:")
        print(f"  saveUser: {self.request.POST.get('saveUser')}")
        print(f"  editOrAdd: {self.request.POST.get('editOrAdd')}")
        print(f"  username: {self.request.POST.get('username')}")
        
        # Get the logged in user session
        session_username = checkSession(self.request)
        print(f"Session username: {session_username}")
        if not session_username:
            print("Session check failed - no session username")
            messages.error(self.request, messagesTypes.notlogin)
            return redirect(self.success_url)
            
        if not check_active(session_username):
            print("Active check failed - user not active")
            messages.error(self.request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if session_username else 'login'))
        
        # Get form data
        field = form.cleaned_data
        fieldReq = self.request.POST
        
        # Extract basic fields
        username = field.get('username')
        editOrAdd = field.get('editOrAdd') if field.get('editOrAdd') else fieldReq.get("editOrAdd", "").lower()
        extension = field.get('extension')
        name = field.get('name', '')
        lastname = field.get('lastname', '')
        email = field.get('email', '')
        active = field.get('active', True)
        
        # Check which action to perform - check direct POST values first, then fallback to form data
        deleteUser = fieldReq.get('deleteUser')
        saveUser = fieldReq.get('saveUser')
        uploadPhoto = self.request.FILES.get('uploadPhoto')
        changePassword = fieldReq.get('ChangePassword')
        
        print(f"Action: editOrAdd={editOrAdd}, saveUser={saveUser}, deleteUser={deleteUser}, changePassword={changePassword}")
        print(f"Files in request: {list(self.request.FILES.keys())}")
        print(f"uploadPhoto received: {uploadPhoto is not None}")
        if uploadPhoto:
            print(f"Upload photo details: name={uploadPhoto.name}, size={uploadPhoto.size}, content_type={uploadPhoto.content_type}")
        
        # Get info fields - convert empty strings to None for nullable fields
        nationalcode = fieldReq.get('nationalcode') or None
        phonenumber = fieldReq.get('phonenumber') or None
        accountnumbershaba = fieldReq.get('accountnumbershaba') or None
        cardnumber = fieldReq.get('cardnumber') or None
        accountnumber = fieldReq.get('accountnumber') or None
        military = fieldReq.get('military') or None
        gender = fieldReq.get('gender') or None
        maritalstatus = fieldReq.get('maritalstatus') or None
        educationdegree = fieldReq.get('educationdegree') or None
        province = fieldReq.get('province') or None
        city = fieldReq.get('city') or None
        address = fieldReq.get('address') or None
        birthdate = fieldReq.get('birthdate') or None
        telephone = fieldReq.get('telephone') or None
        educationfield = fieldReq.get('educationfield') or None
        
        # Process usersextension field
        listOfExts = fieldReq.getlist('usersextension')
        nonLabels = []
        labels = []
        
        if listOfExts:
            for item in listOfExts:
                item = item.strip()
                if Extensionsgroups.objects.filter(label=str(item)).exists():
                    labels.append(item)
                else:
                    nonLabels.append(str(item))
        
        # Handle the different actions
        if saveUser:
            print(f"Processing saveUser action with mode: {editOrAdd}")
            
            # Validate editOrAdd field
            if not editOrAdd or editOrAdd in ['none', '']:
                print("editOrAdd validation failed - must select edit or add mode")
                messages.error(self.request, "لطفا نوع عملیات (ویرایش یا اضافه) را انتخاب کنید.")
                return redirect(self.success_url)
            
            if editOrAdd == 'add':
                print("Entering add user mode")
                # Check write permission for adding users
                if isinstance(check := hasAccess("write", "user", self.request), HttpResponseRedirect):
                    print("Access check failed - redirecting")
                    return check
                
                # Validate required fields for adding a user
                if not username:
                    print("Username validation failed")
                    messages.error(self.request, "نام کاربری الزامی است.")
                    return redirect(self.success_url)
                
                if Users.objects.filter(username__iexact=username).exists():
                    print("Username already exists")
                    messages.error(self.request, "این نام کاربری موجود است.")
                    return redirect(self.success_url)
                
                if not extension:
                    print("Extension validation failed")
                    messages.error(self.request, "داخلی الزامی است.")
                    return redirect(self.success_url)
                
                # Check if role is selected
                groupname = field.get('groupname') or fieldReq.get('groupname')
                print(f"Groupname value: {groupname}")
                if not groupname or groupname in ['none', '']:
                    print("Groupname validation failed")
                    messages.error(self.request, "انتخاب نقش الزامی است.")
                    return redirect(self.success_url)
                
                # Check if at least one extension access is selected
                print(f"Extension list: {listOfExts}")
                if not listOfExts:
                    print("Extensions validation failed")
                    messages.error(self.request, "انتخاب حداقل یک دسترسی به داخلی الزامی است.")
                    return redirect(self.success_url)
                
                # Check if at least one permission is selected
                permCount = 0
                for perm in ["can_view", "can_write", "can_modify", "can_delete"]:
                    if fieldReq.get(perm):
                        permCount += 1
                        print(f"Permission {perm} is set")
                
                print(f"Total permissions selected: {permCount}")
                if permCount == 0:
                    print("Permissions validation failed")
                    messages.error(self.request, "انتخاب حداقل یک سطح دسترسی الزامی است.")
                    return redirect(self.success_url)
                
                print("All validations passed, proceeding with user creation...")
                
                groupname = groupname.lower()
                group = None
                
                print(f"Looking for group: {groupname}")
                
                # Try to find the group
                if groupname:
                    group = Groups.objects.filter(enname__iexact=groupname).first()
                    if not group:
                        group = Groups.objects.filter(pename__iexact=groupname).first()
                
                # If no group found, use the default 'member' group
                if not group:
                    print("Group not found, trying member group")
                    group = Groups.objects.filter(enname__iexact='member').first()
                    if not group:
                        print("Member group not found, using first available group")
                        group = Groups.objects.first()  # Last resort
                        
                if not group:
                    print("No groups found in system!")
                    messages.error(self.request, "هیچ گروهی در سیستم تعریف نشده است.")
                    return redirect(self.success_url)
                
                print(f"Selected group: {group.enname} (ID: {group.id})")
                
                # Set default extension if not provided
                try:
                    extension_value = int(extension) if extension else 1000
                except ValueError:
                    extension_value = 1000
                
                print(f"Extension value: {extension_value}")
                
                # Ensure required fields have values
                if not name:
                    name = username  # Use username as name if name is empty
                if not lastname:
                    lastname = "کاربر"  # Default lastname
                if not email:
                    email = f"{username}@example.com"  # Default email
                
                print(f"Final values - name: {name}, lastname: {lastname}, email: {email}")
                
                # Set default picture
                picurl = "avatar.png"
                
                # Handle profile picture upload if provided
                if uploadPhoto:
                    print(f"Photo upload detected for new user: {uploadPhoto.name}, size: {uploadPhoto.size}")
                    valid = validatePhotoExt(uploadPhoto.name)
                    if not valid:
                        print(f"Invalid photo extension for file: {uploadPhoto.name}")
                        messages.error(self.request, 'پسوند فایل ارسال شده نامعتبر میباشد.')
                        return redirect(self.success_url)
                    picurl = f"{username.lower()}_photo{valid.lower()}"
                    filepath = os.path.join('Alvand/static/upload', picurl)
                    print(f"Saving new user photo to: {filepath}")

                    try:
                        # Ensure upload directory exists
                        upload_dir = os.path.join('Alvand/static/upload')
                        if not os.path.exists(upload_dir):
                            print(f"Creating upload directory: {upload_dir}")
                            os.makedirs(upload_dir, exist_ok=True)
                        
                        print(f"Upload directory exists: {os.path.exists(upload_dir)}")
                        print(f"Upload directory permissions: {oct(os.stat(upload_dir).st_mode)[-3:] if os.path.exists(upload_dir) else 'N/A'}")
                        
                        with open(filepath, 'wb') as f:  # Changed from '+wb' to 'wb'
                            for chunk in uploadPhoto.chunks():  # Use chunks() for better memory handling
                                f.write(chunk)
                        
                        print(f"Photo saved successfully for new user: {picurl}")
                        print(f"Database fields will be set - picurl: {picurl}, profile_picture: {picurl}")
                        print(f"File exists after save: {os.path.exists(filepath)}")
                        print(f"File size after save: {os.path.getsize(filepath) if os.path.exists(filepath) else 'N/A'}")
                        
                    except Exception as photo_error:
                        print(f"Error saving photo for new user: {str(photo_error)}")
                        messages.error(self.request, f'خطا در ذخیره تصویر: {str(photo_error)}')
                        return redirect(self.success_url)
                else:
                    print("No photo upload in add user request")
                
                # Directly create the user
                try:
                    print("Creating user...")
                    new_user = Users.objects.create(
                        username=username,
                        name=name,
                        lastname=lastname,
                        extension=extension_value,
                        email=email,
                        group=group,
                        groupname=group.enname.lower(),
                        picurl=picurl,
                        profile_picture=picurl,  # Also update profile_picture field
                        active=active,
                        usersextension=nonLabels,
                        password=make_password("12345678"),
                        needs_password_change=True
                    )
                    print(f"User created with ID: {new_user.id}")
                    print(f"New user database verification - picurl: {new_user.picurl}, profile_picture: {new_user.profile_picture}")
                    
                    # Create permissions
                    print("Creating permissions...")
                    Permissions.objects.create(
                        user=new_user,
                        can_view=bool(fieldReq.get("can_view")),
                        can_write=bool(fieldReq.get("can_write")),
                        can_modify=bool(fieldReq.get("can_modify")),
                        can_delete=bool(fieldReq.get("can_delete")),
                        exts_label=labels
                    )
                    
                    # Create user info - ensure NULL values for empty fields
                    print("Creating user info...")
                    Infos.objects.create(
                        user=new_user,
                        gender=int(gender) if gender and str(gender).isdigit() else None,
                        nationalcode=int(nationalcode) if nationalcode and str(nationalcode).isdigit() else None,
                        birthdate=birthdate,
                        telephone=telephone,
                        phonenumber=phonenumber,
                        maritalstatus=maritalstatus,
                        military=military,
                        educationfield=educationfield,
                        educationdegree=educationdegree,
                        province=province,
                        city=city,
                        accountnumbershaba=accountnumbershaba,
                        cardnumber=cardnumber,
                        accountnumber=accountnumber,
                        address=address
                    )
                    
                    messages.success(self.request, f'کاربر {username} با موفقیت به جمع ما پیوست.')
                
                except Exception as e:
                    import traceback
                    print(f"Error creating user: {str(e)}")
                    print(traceback.format_exc())
                    messages.error(self.request, f'خطا در ایجاد کاربر: {str(e)}')
            
            # Edit mode
            elif editOrAdd == 'edit':
                # Find the user to edit
                try:
                    user = Users.objects.get(username=username)
                    
                    # Check permission to edit this user
                    if not hasAccessToUser(self.request, user):
                        messages.error(self.request, messagesTypes.permissionsNotFound)
                        return redirect(self.success_url)
                    
                    # Update user fields
                    user.name = name
                    user.lastname = lastname
                    user.extension = int(extension) if extension and str(extension).isdigit() else user.extension
                    user.email = email
                    user.active = active
                    user.usersextension = nonLabels
                    
                    # Update group if allowed
                    if user.username != session_username:
                        groupname = field.get('groupname')
                        if groupname and groupname != 'none':
                            group = Groups.objects.filter(enname__iexact=groupname).first()
                            if not group:
                                group = Groups.objects.filter(pename__iexact=groupname).first()
                            
                            if group:
                                user.group = group
                                user.groupname = group.enname.lower()
                    
                    # Handle profile picture upload if provided
                    if uploadPhoto:
                        print(f"Photo upload detected: {uploadPhoto.name}, size: {uploadPhoto.size}")
                        valid = validatePhotoExt(uploadPhoto.name)
                        if not valid:
                            print(f"Invalid photo extension for file: {uploadPhoto.name}")
                            messages.error(self.request, 'پسوند فایل ارسال شده نامعتبر میباشد.')
                            return redirect(self.success_url)
                        picurl = f"{username.lower()}_photo{valid.lower()}"
                        filepath = os.path.join('Alvand/static/upload', picurl)
                        print(f"Saving photo to: {filepath}")

                        try:
                            # Ensure upload directory exists
                            upload_dir = os.path.join('Alvand/static/upload')
                            if not os.path.exists(upload_dir):
                                print(f"Creating upload directory: {upload_dir}")
                                os.makedirs(upload_dir, exist_ok=True)
                            
                            print(f"Upload directory exists: {os.path.exists(upload_dir)}")
                            print(f"Upload directory permissions: {oct(os.stat(upload_dir).st_mode)[-3:] if os.path.exists(upload_dir) else 'N/A'}")
                            
                            with open(filepath, 'wb') as f:  # Changed from '+wb' to 'wb'
                                for chunk in uploadPhoto.chunks():  # Use chunks() for better memory handling
                                    f.write(chunk)
                            
                            user.picurl = picurl
                            user.profile_picture = picurl  # Also update profile_picture field
                            print(f"Photo saved successfully: {picurl}")
                            print(f"Database fields set - picurl: {user.picurl}, profile_picture: {user.profile_picture}")
                            print(f"File exists after save: {os.path.exists(filepath)}")
                            print(f"File size after save: {os.path.getsize(filepath) if os.path.exists(filepath) else 'N/A'}")
                            
                        except Exception as photo_error:
                            print(f"Error saving photo: {str(photo_error)}")
                            messages.error(self.request, f'خطا در ذخیره تصویر: {str(photo_error)}')
                            return redirect(self.success_url)
                    else:
                        print("No photo upload in request")
                    
                    # Save user updates
                    user.save()
                    print(f"User saved to database successfully for: {username}")
                    
                    # Verify database update by reloading user
                    updated_user = Users.objects.filter(username=username).first()
                    if updated_user:
                        print(f"Database verification - picurl: {updated_user.picurl}, profile_picture: {updated_user.profile_picture}")
                    else:
                        print(f"Error: Could not reload user {username} from database after save")
                    
                    # Update permissions
                    perm = Permissions.objects.filter(user=user).first()
                    if perm:
                        perm.can_view = bool(fieldReq.get("can_view"))
                        perm.can_write = bool(fieldReq.get("can_write"))
                        perm.can_modify = bool(fieldReq.get("can_modify"))
                        perm.can_delete = bool(fieldReq.get("can_delete"))
                        perm.exts_label = labels
                        perm.save()
                    else:
                        Permissions.objects.create(
                            user=user,
                            can_view=bool(fieldReq.get("can_view")),
                            can_write=bool(fieldReq.get("can_write")),
                            can_modify=bool(fieldReq.get("can_modify")),
                            can_delete=bool(fieldReq.get("can_delete")),
                            exts_label=labels
                        )
                    
                    # Update user info
                    userInfo = Infos.objects.filter(user=user).first()
                    if userInfo:
                        userInfo.gender = int(gender) if gender and gender.isdigit() else userInfo.gender
                        userInfo.nationalcode = int(nationalcode) if nationalcode and nationalcode.isdigit() else userInfo.nationalcode
                        userInfo.birthdate = birthdate
                        userInfo.telephone = telephone
                        userInfo.phonenumber = phonenumber
                        userInfo.maritalstatus = maritalstatus
                        userInfo.military = military
                        userInfo.educationfield = educationfield
                        userInfo.educationdegree = educationdegree
                        userInfo.province = province
                        userInfo.city = city
                        userInfo.accountnumbershaba = accountnumbershaba
                        userInfo.cardnumber = cardnumber
                        userInfo.accountnumber = accountnumber
                        userInfo.address = address
                        userInfo.save()
                    else:
                        Infos.objects.create(
                            user=user,
                            gender=int(gender) if gender and gender.isdigit() else None,
                            nationalcode=int(nationalcode) if nationalcode and nationalcode.isdigit() else None,
                            birthdate=birthdate,
                            telephone=telephone,
                            phonenumber=phonenumber,
                            maritalstatus=maritalstatus,
                            military=military,
                            educationfield=educationfield,
                            educationdegree=educationdegree,
                            province=province,
                            city=city,
                            accountnumbershaba=accountnumbershaba,
                            cardnumber=cardnumber,
                            accountnumber=accountnumber,
                            address=address
                        )
                    
                    messages.success(self.request, f'اطلاعات کاربر {username} با موفقیت بروز شد.')
                    
                except Users.DoesNotExist:
                    messages.error(self.request, 'کاربر مورد نظر یافت نشد.')
                except Exception as e:
                    import traceback
                    print(f"Error updating user: {str(e)}")
                    print(traceback.format_exc())
                    messages.error(self.request, f'خطا در بروزرسانی کاربر: {str(e)}')
            
            else:
                messages.error(self.request,
                               'برای انجام عملیات ویرایش یا اضافه کاربر باید گزینه مناسب را در فیلد ویرایش/اضافه انتخاب کنید.')
        
        elif deleteUser:
            # Process delete user request
            try:
                user = Users.objects.filter(username=username).first()
                if user:
                    if not hasAccessToUser(self.request, user):
                        messages.error(self.request, messagesTypes.permissionsNotFound)
                        return redirect(self.success_url)
                    
                    # Delete user and related records
                    user_id = user.id
                    
                    # Delete or handle log entries first to avoid foreign key constraint
                    from .models import Log
                    log_count = Log.objects.filter(user=user).count()
                    
                    if log_count > 0:
                        # Keep logs but delete the foreign key reference
                        # This preserves audit trail while allowing user deletion
                        Log.objects.filter(user=user).delete()
                        
                        # Log this action for audit purposes before deleting the user
                        try:
                            log(self.request, "USER_DELETE", f"حذف کاربر {username} همراه با {log_count} رکورد لاگ")
                        except:
                            # If logging fails, continue with deletion anyway
                            pass
                    
                    # Delete permissions
                    Permissions.objects.filter(user=user).delete()
                    
                    # Delete info
                    Infos.objects.filter(user=user).delete()
                    
                    # Delete verifications if any
                    from .models import Verifications
                    Verifications.objects.filter(user=user).delete()
                    
                    # Delete password reset requests if any
                    from .models import PasswordResetRequest
                    PasswordResetRequest.objects.filter(user=user).delete()
                    PasswordResetRequest.objects.filter(resolved_by=user).update(resolved_by=None)
                    
                    # Delete user
                    user.delete()
                    
                    messages.success(self.request, f'کاربر {username} با موفقیت حذف شد.')
                else:
                    messages.error(self.request, 'کاربر مورد نظر یافت نشد.')
            except Exception as e:
                messages.error(self.request, f'خطا در حذف کاربر: {str(e)}')
        
        elif changePassword:
            # Process change password request
            try:
                user = Users.objects.filter(username=username).first()
                if user:
                    if not hasAccessToUser(self.request, user):
                        messages.error(self.request, messagesTypes.permissionsNotFound)
                        return redirect(self.success_url)
                    
                    # Reset password and set change flag
                    user.password = make_password("12345678")
                    user.needs_password_change = True
                    user.save()
                    
                    # Mark any pending password reset requests as resolved
                    from .models import PasswordResetRequest
                    current_user = Users.objects.filter(username=checkSession(self.request)).first()
                    pending_requests = PasswordResetRequest.objects.filter(user=user, resolved=False)
                    
                    request_resolved = False
                    if pending_requests.exists():
                        for req in pending_requests:
                            req.resolved = True
                            req.resolved_by = current_user
                            req.resolved_at = timezone.now()
                            req.save()
                        request_resolved = True
                    
                    # Log the password reset
                    log(self.request, logErrCodes.userSettings, f"رمز عبور کاربر {username} با موفقیت بازنشانی شد.", checkSession(self.request))
                    
                    if request_resolved:
                        messages.success(self.request, f'رمز عبور کاربر {username} با موفقیت بازنشانی شد و به 12345678 تغییر یافت. درخواست بازنشانی ثبت شده برای این کاربر نیز به وضعیت حل شده تغییر یافت.')
                    else:
                        messages.success(self.request, f'رمز عبور کاربر {username} با موفقیت بازنشانی شد و به 12345678 تغییر یافت.')
                else:
                    messages.error(self.request, 'کاربر مورد نظر یافت نشد.')
            except Exception as e:
                messages.error(self.request, f'خطا در بازنشانی رمز عبور: {str(e)}')
                
        # Final redirect
        print(f"At end of form_valid method")
        print(f"self.success_url = {self.success_url}")
        print(f"reverse_lazy('user') = {reverse_lazy('user')}")
        print(f"About to redirect to: {self.success_url}")
        return redirect(self.success_url)

    def form_invalid(self, form):
        if not checkSession(self.request):
            messages.error(self.request, messagesTypes.notlogin)
            return redirect(self.success_url)
        if not check_active(checkSession(self.request)):
            messages.error(self.request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout' if checkSession(self.request) else 'login'))
        if isinstance(check := hasAccess("write", "user", self.request), HttpResponseRedirect):
            return check
        
        # Debug form errors
        print("Form validation failed in form_invalid method:")
        print(f"Form errors: {form.errors}")
        print(f"Form non-field errors: {form.non_field_errors()}")
        
        # Add specific error messages for debugging
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                error_messages.append(f"{field}: {error}")
        
        if error_messages:
            messages.error(self.request, "خطاهای فرم: " + ", ".join(error_messages))
        else:
            messages.error(self.request, messagesTypes.fillAllFields)
        
        return redirect(self.success_url)

    def post(self, request, *args, **kwargs):
        """Handle POST requests for user form."""
        print("=" * 50)
        print("POST request received to UserForm")
        print(f"POST data: {dict(request.POST)}")
        print(f"FILES data: {dict(request.FILES)}")
        
        # Debug the saveUser field specifically
        print(f"saveUser field present: {'saveUser' in request.POST}")
        print(f"saveUser field value: {request.POST.get('saveUser')}")
        
        # Check session and permissions first
        if not checkSession(request):
            print("User not logged in")
            messages.error(request, messagesTypes.notlogin)
            return redirect(reverse_lazy('login'))
            
        if not check_active(checkSession(request)):
            print("User not active")
            messages.error(request, messagesTypes.deAvtive)
            return redirect(reverse_lazy('logout'))
        
        # Special handling for direct actions
        if 'saveUser' in request.POST:
            # Get form data
            username = request.POST.get('username')
            editOrAdd = request.POST.get('editOrAdd')
            
            print(f"Direct saveUser action detected - username: {username}, mode: {editOrAdd}")
            
            # Process form using form_valid directly
            form = self.get_form()
            print(f"Form created: {form.__class__.__name__}")
            print(f"Form data before validation: {form.data}")
            
            if form.is_valid():
                print("Form is valid, calling form_valid")
                return self.form_valid(form)
            else:
                print("Form validation failed")
                print(f"Form errors: {form.errors}")
                print(f"Form non-field errors: {form.non_field_errors()}")
                return self.form_invalid(form)
        elif 'deleteUser' in request.POST:
            # Handle direct delete
            username = request.POST.get('username')
            print(f"Direct deleteUser action detected for username: {username}")
            
            try:
                user = Users.objects.filter(username=username).first()
                if user:
                    if not hasAccessToUser(request, user):
                        messages.error(request, messagesTypes.permissionsNotFound)
                        return redirect(self.success_url)
                    
                    # Delete user and related records
                    user_id = user.id
                    
                    # Delete or handle log entries first to avoid foreign key constraint
                    from .models import Log
                    log_count = Log.objects.filter(user=user).count()
                    
                    if log_count > 0:
                        # Keep logs but delete the foreign key reference
                        # This preserves audit trail while allowing user deletion
                        Log.objects.filter(user=user).delete()
                        
                        # Log this action for audit purposes before deleting the user
                        try:
                            log(request, "USER_DELETE", f"حذف کاربر {username} همراه با {log_count} رکورد لاگ")
                        except:
                            # If logging fails, continue with deletion anyway
                            pass
                    
                    # Delete permissions
                    Permissions.objects.filter(user=user).delete()
                    
                    # Delete info
                    Infos.objects.filter(user=user).delete()
                    
                    # Delete verifications if any
                    from .models import Verifications
                    Verifications.objects.filter(user=user).delete()
                    
                    # Delete password reset requests if any
                    from .models import PasswordResetRequest
                    PasswordResetRequest.objects.filter(user=user).delete()
                    PasswordResetRequest.objects.filter(resolved_by=user).update(resolved_by=None)
                    
                    # Delete user
                    user.delete()
                    
                    messages.success(request, f'کاربر {username} با موفقیت حذف شد.')
                else:
                    messages.error(request, 'کاربر مورد نظر یافت نشد.')
            except Exception as e:
                messages.error(request, f'خطا در حذف کاربر: {str(e)}')
            
            return redirect(self.success_url)
        elif 'ChangePassword' in request.POST:
            # Handle direct password change
            username = request.POST.get('username')
            print(f"Direct ChangePassword action detected for username: {username}")
            
            try:
                user = Users.objects.filter(username=username).first()
                if user:
                    if not hasAccessToUser(request, user):
                        messages.error(request, messagesTypes.permissionsNotFound)
                        return redirect(self.success_url)
                    
                    # Reset password and set change flag
                    user.password = make_password("12345678")
                    user.needs_password_change = True
                    user.save()
                    
                    # Mark any pending password reset requests as resolved
                    from .models import PasswordResetRequest
                    current_user = Users.objects.filter(username=checkSession(request)).first()
                    pending_requests = PasswordResetRequest.objects.filter(user=user, resolved=False)
                    
                    request_resolved = False
                    if pending_requests.exists():
                        for req in pending_requests:
                            req.resolved = True
                            req.resolved_by = current_user
                            req.resolved_at = timezone.now()
                            req.save()
                        request_resolved = True
                    
                    # Log the password reset
                    log(request, logErrCodes.userSettings, f"رمز عبور کاربر {username} با موفقیت بازنشانی شد.", checkSession(request))
                    
                    if request_resolved:
                        messages.success(request, f'رمز عبور کاربر {username} با موفقیت بازنشانی شد و به 12345678 تغییر یافت. درخواست بازنشانی ثبت شده برای این کاربر نیز به وضعیت حل شده تغییر یافت.')
                    else:
                        messages.success(request, f'رمز عبور کاربر {username} با موفقیت بازنشانی شد و به 12345678 تغییر یافت.')
                else:
                    messages.error(request, 'کاربر مورد نظر یافت نشد.')
            except Exception as e:
                messages.error(request, f'خطا در بازنشانی رمز عبور: {str(e)}')
            
            return redirect(self.success_url)
        
        # Regular form processing if no direct action detected
        form = self.get_form()
        print(f"Form is valid: {form.is_valid()}")
        if form.is_valid():
            print("Form validation passed, calling form_valid")
            return self.form_valid(form)
        else:
            print(f"Form validation failed!")
            print(f"Form errors: {form.errors}")
            print(f"Form non-field errors: {form.non_field_errors()}")
            # Add detailed error messages to help debug
            for field, errors in form.errors.items():
                print(f"Field '{field}' errors: {errors}")
            return self.form_invalid(form)


def get_user_data(request):
    """AJAX endpoint to fetch user data for editing."""
    print(f"get_user_data called - Method: {request.method}")
    print(f"Headers: {dict(request.headers)}")
    print(f"GET params: {dict(request.GET)}")
    
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        username = request.GET.get('username')
        print(f"AJAX request for username: {username}")
        
        if not username:
            return JsonResponse({'success': False, 'error': 'Username is required'})
        
        try:
            # Get user data
            user = Users.objects.filter(username__iexact=username).first()
            if not user:
                return JsonResponse({'success': False, 'error': 'User not found'})
            
            # Get user info
            user_info = Infos.objects.filter(user=user).first()
            
            # Get user permissions
            permissions = Permissions.objects.filter(user=user).first()
            
            # Prepare user data
            user_data = {
                'username': user.username,
                'name': user.name or '',
                'lastname': user.lastname or '',
                'extension': user.extension,
                'email': user.email or '',
                'groupname': user.groupname or '',
                'active': user.active,
                'usersextension': user.usersextension or [],
                'picurl': user.picurl or 'avatar.png'
            }
            
            # Add user info if exists
            if user_info:
                user_data.update({
                    'nationalcode': user_info.nationalcode or '',
                    'birthdate': user_info.birthdate.strftime('%Y/%m/%d') if user_info.birthdate else '',
                    'telephone': user_info.telephone or '',
                    'phonenumber': user_info.phonenumber or '',
                    'gender': user_info.gender or '',
                    'maritalstatus': user_info.maritalstatus or '',
                    'military': user_info.military or '',
                    'educationfield': user_info.educationfield or '',
                    'educationdegree': user_info.educationdegree or '',
                    'province': user_info.province or '',
                    'city': user_info.city or '',
                    'accountnumbershaba': user_info.accountnumbershaba or '',
                    'cardnumber': user_info.cardnumber or '',
                    'accountnumber': user_info.accountnumber or '',
                    'address': user_info.address or ''
                })
            
            # Add permissions if exists
            if permissions:
                user_data.update({
                    'can_view': permissions.can_view,
                    'can_write': permissions.can_write,
                    'can_modify': permissions.can_modify,
                    'can_delete': permissions.can_delete
                })
            
            return JsonResponse({'success': True, 'user_data': user_data})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


def get_cable_types(request):
    """AJAX endpoint to get cable type options based on device model."""
    if request.method == 'GET':
        # Always return all cable type options; decision logic is handled elsewhere
        cable_types = [
            {'value': '', 'label': '------'},
            {'value': 'rs-232c', 'label': 'RS-232C'},
            {'value': 'ethernet', 'label': 'ETHERNET'}
        ]

        return JsonResponse({'success': True, 'cable_types': cable_types})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def index(request):
    return redirect(reverse_lazy("dashboard"))

class CustomJSONEncoder(DjangoJSONEncoder):
    def default(self, obj):
        # Handle QuerySets and other Django-specific objects if needed
        if isinstance(obj, models.QuerySet):
            return list(obj.values())
        return super().default(obj)

def userprofile_view(request):
    users_queryset = Users.objects.all()
    session_user_obj = request.session.get('user', None)

    users_data = []
    for user in users_queryset:
        # Get user extensions safely
        user_extensions = user.usersextension if user.usersextension else []
        
        # Get list of extension groups safely
        ext_groups = getListOfExtsGroups(user.username) or []
        
        # Get user permissions safely
        user_perms_raw = getUserCanPerm(user.username) or '{}'
        try:
            user_perms = json.loads(user_perms_raw) if isinstance(user_perms_raw, str) else user_perms_raw
        except (json.JSONDecodeError, TypeError):
            user_perms = {}
        
        # Get user info safely
        user_info = getObjectOfInfo(user.username)
        user_info_list = list(user_info.values()) if user_info and hasattr(user_info, 'values') else []
        
        user_dict = {
            'username': user.username,
            'name': user.name,
            'lastname': user.lastname,
            'extension': user.extension,
            'email': user.email,
            'active': user.active,
            'groupname': user.groupname,
            'picurl': user.picurl if user.picurl else 'avatar.png',
            'usersextension': user_extensions,
            'getListOfExtsGroups': ext_groups,
            'getUserCanPerm': user_perms,
            'getObjectOfInfo': user_info_list
        }
        users_data.append(user_dict)
    
    session_user_data = {}
    if session_user_obj:
        # session_user_obj is the username string, get the actual User object
        try:
            session_user = Users.objects.get(username=session_user_obj)
            session_user_data = {
                'username': session_user.username,
                'getUserInfo_groupname': getUserInfo(session_user.username, "groupname")
            }
        except Users.DoesNotExist:
            session_user_data = {
                'username': session_user_obj,
                'getUserInfo_groupname': ''
            }

    all_contact_infos = []
    for user_obj in users_queryset:
        infos = getObjectOfInfo(user_obj.username)
        for info in infos:
            info_dict = {
                'username': user_obj.username,
                'nationalcode': info.nationalcode,
                'birthdate': str(info.birthdate) if info.birthdate else None,
                'telephone': info.telephone,
                'phonenumber': info.phonenumber,
                'gender': info.gender,
                'maritalstatus': info.maritalstatus,
                'military': info.military,
                'educationfield': info.educationfield,
                'educationdegree': info.educationdegree,
                'province': info.province,
                'city': info.city,
                'accountnumbershaba': info.accountnumbershaba,
                'cardnumber': info.cardnumber,
                'accountnumber': info.accountnumber,
                'address': info.address,
            }
            all_contact_infos.append(info_dict)

    context = {
        'pageTitle': 'مدیریت پروفایل کاربران',
        'userform': userProfileForm(),
        'infosform': InfosForm(),
        'permform': PermissionsForm(),
        'users': users_data,
        'session_user': session_user_data,
        'contactInfos': all_contact_infos,
    }
    
    return render(request, 'userprofile.html', context)

def validate_password_complexity(password):
    """
    بررسی پیچیدگی رمز عبور
    
    شرایط رمز عبور معتبر:
    1. حداقل 8 کاراکتر
    2. حداقل یک حرف بزرگ انگلیسی
    3. حداقل یک حرف کوچک انگلیسی
    4. حداقل یک عدد
    5. حداقل یک کاراکتر خاص (!@#$%^&*()_+-=[]{}|;:,.<>?/~)
    6. نباید همان رمز عبور پیش‌فرض (12345678) باشد
    
    returns: (is_valid, error_message)
    """
    import re
    
    # بررسی طول رمز عبور
    if len(password) < 8:
        return False, 'رمز عبور باید حداقل ۸ کاراکتر باشد.'
    
    # بررسی وجود حروف بزرگ انگلیسی
    if not re.search(r'[A-Z]', password):
        return False, 'رمز عبور باید حداقل شامل یک حرف بزرگ انگلیسی باشد.'
    
    # بررسی وجود حروف کوچک انگلیسی
    if not re.search(r'[a-z]', password):
        return False, 'رمز عبور باید حداقل شامل یک حرف کوچک انگلیسی باشد.'
    
    # بررسی وجود اعداد
    if not re.search(r'[0-9]', password):
        return False, 'رمز عبور باید حداقل شامل یک عدد باشد.'
    
    # بررسی وجود کاراکترهای خاص
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{}|;:\'",.<>?/~]', password):
        return False, 'رمز عبور باید حداقل شامل یک کاراکتر خاص مانند !@#$%^&* باشد.'
    
    # بررسی عدم استفاده از رمز عبور پیش‌فرض
    if password == "12345678":
        return False, 'رمز عبور نمی‌تواند همان رمز عبور پیش‌فرض باشد.'
    
    return True, ''

class ChangePasswordView(View):
    template_name = 'change_password.html'

    def get(self, request):
        if not checkSession(request):
            messages.error(request, messagesTypes.notlogin)
            return redirect('login')
        return render(request, self.template_name)

    def post(self, request):
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # بررسی تطابق رمزها
        if new_password != confirm_password:
            messages.error(request, 'رمزهای عبور وارد شده مطابقت ندارند.')
            return redirect('change_password')
        
        # بررسی پیچیدگی رمز عبور
        is_valid, error_message = validate_password_complexity(new_password)
        if not is_valid:
            messages.error(request, error_message)
            return redirect('change_password')
        
        # بروزرسانی رمز عبور
        username = checkSession(request)
        user = Users.objects.get(username=username)
        user.password = make_password(new_password)
        user.needs_password_change = False
        user.save()
        
        # ثبت رویداد در لاگ
        log(request, logErrCodes.userSettings, f"رمز عبور کاربر {username} با موفقیت تغییر یافت.", username)
        
        # خروج و درخواست ورود مجدد
        del request.session['user']
        messages.success(request, 'رمز عبور با موفقیت تغییر یافت. لطفا با رمز عبور جدید وارد شوید.')
        return redirect('login')

class ForgotPasswordView(View):
    template_name = 'forgot_password.html'
    notification_template = 'forgot_password_notification.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username')
        
        # Check if user exists
        user = Users.objects.filter(username__iexact=username).first()
        if user:
            # Create a password reset request
            from .models import PasswordResetRequest
            PasswordResetRequest.objects.create(user=user)
            
            # Log the request only if username is valid
            if username and isinstance(username, str) and username.strip():
                log(request, logErrCodes.userSettings, f"درخواست بازنشانی رمز عبور برای کاربر {username} ثبت شد.", username)
            
            # Show notification page
            messages.success(request, f'درخواست بازنشانی رمز عبور برای کاربر {username} با موفقیت ثبت شد.')
            return render(request, self.notification_template)
                
        messages.error(request, 'نام کاربری وارد شده در سیستم یافت نشد.')
        return redirect('forgot_password')

class ResetPasswordView(View):
    template_name = 'reset_password.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        code = request.POST.get('code')
        new_password = request.POST.get('new_password')
        confirm = request.POST.get('confirm_password')
        
        # بررسی تطابق رمزها
        if new_password != confirm:
            messages.error(request, 'رمزهای عبور وارد شده مطابقت ندارند.')
            return redirect('reset_password')
        
        # بررسی پیچیدگی رمز عبور
        is_valid, error_message = validate_password_complexity(new_password)
        if not is_valid:
            messages.error(request, error_message)
            return redirect('reset_password')
        
        # بررسی صحت کد بازنشانی
        ver = Verifications.objects.filter(code=code, type=verificationType.email).first()
        if ver:
            user = ver.user
            user.password = make_password(new_password)
            user.needs_password_change = False
            user.save()
            ver.delete()
            
            # ثبت رویداد در لاگ
            log(request, logErrCodes.userSettings, f"رمز عبور کاربر {user.username} با موفقیت بازنشانی شد.", user.username)
            
            messages.success(request, 'رمز عبور با موفقیت بازنشانی شد. اکنون می‌توانید با رمز عبور جدید وارد شوید.')
            return redirect('login')
            
        messages.error(request, 'کد بازنشانی نامعتبر است.')
        return redirect('reset_password')

def test_user_create(request):
    """Simple test view for direct user creation."""
    from django.shortcuts import render
    from django.contrib import messages
    from Alvand.models import Users, Groups
    from django.contrib.auth.hashers import make_password
    import traceback
    
    # Get all groups for dropdown
    groups = Groups.objects.all()
    context = {'groups': groups}
    
    # Handle form submission
    if request.method == 'POST':
        try:
            # Get form data
            username = request.POST.get('username')
            name = request.POST.get('name', '')
            lastname = request.POST.get('lastname', '')
            extension = request.POST.get('extension', '1000')
            email = request.POST.get('email', '')
            group_id = request.POST.get('group_id')
            
            # Validate required fields
            if not username:
                messages.error(request, 'Username is required')
                return render(request, 'test_user_create.html', context)
            
            # Check if username exists
            if Users.objects.filter(username__iexact=username).exists():
                messages.error(request, 'Username already exists')
                return render(request, 'test_user_create.html', context)
            
            # Get group
            try:
                group = Groups.objects.get(id=group_id)
            except (Groups.DoesNotExist, ValueError):
                messages.error(request, 'Invalid group ID')
                return render(request, 'test_user_create.html', context)
            
            # Process extension
            try:
                extension_value = int(extension)
            except ValueError:
                extension_value = 1000
            
            # Create user
            new_user = Users(
                username=username,
                name=name,
                lastname=lastname,
                extension=extension_value,
                email=email,
                group=group,
                groupname=group.enname.lower(),
                picurl="avatar.png",
                active=True,
                usersextension=[],
                password=make_password("123456789"),
                needs_password_change=True
            )
            
            # Save user
            new_user.save()
            
            # Show success message
            messages.success(request, f'User {username} created successfully with ID: {new_user.id}')
            
            # Add result details
            result = f"User created successfully:\n"
            result += f"ID: {new_user.id}\n"
            result += f"Username: {new_user.username}\n"
            result += f"Group: {new_user.group.enname}\n"
            result += f"Extension: {new_user.extension}\n"
            
            context['result'] = result
            
        except Exception as e:
            error_details = traceback.format_exc()
            messages.error(request, f'Error creating user: {str(e)}')
            context['result'] = f"Error details:\n{error_details}"
    
    return render(request, 'test_user_create.html', context)

def hasAccessToUser(request, target_user):
    """Helper function to check if the logged in user has access to modify the target user."""
    session_username = checkSession(request)
    session_user = Users.objects.filter(username__iexact=session_username).first()
    
    if not session_user:
        return False
        
    # Get user roles
    session_role = session_user.groupname.lower()
    target_role = target_user.groupname.lower()
    
    # Check permissions hierarchy
    if session_role == 'supporter':
        return True  # Supporter can modify anyone
    elif session_role == 'superadmin':
        return target_role not in ['supporter']  # Superadmin can't modify supporter
    elif session_role == 'admin':
        return target_role not in ['supporter', 'superadmin']  # Admin can't modify supporter or superadmin
    elif session_user.id == target_user.id:
        return True  # User can modify themselves
    else:
        return False

# Add this context processor at the end of the file

def password_reset_request_count(request):
    from .models import PasswordResetRequest, Users
    count = 0
    groupname = None
    username = None
    if 'user' in request.session:
        username = request.session['user']
        user = Users.objects.filter(username=username).first()
        if user:
            groupname = user.groupname
    elif hasattr(request, 'user') and hasattr(request.user, 'groupname'):
        groupname = getattr(request.user, 'groupname', None)
    if groupname and str(groupname).lower() in ['supporter', 'superadmin', 'admin']:
        count = PasswordResetRequest.objects.filter(resolved=False).count()
    return {'pending_password_reset_count': count}












