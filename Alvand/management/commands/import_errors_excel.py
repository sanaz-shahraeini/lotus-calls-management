import os
from django.core.management.base import BaseCommand
from django.db import transaction
from Alvand.models import Errors

class Command(BaseCommand):
    help = 'Import Errors from an Excel (.xlsx) file into the Errors model'

    def add_arguments(self, parser):
        parser.add_argument('-f', '--file', dest='file', default='Panasonic SMDR List of Errors and Solutions.xlsx',
                            help='Path to the Excel .xlsx file')
        parser.add_argument('-s', '--sheet', dest='sheet', default='English',
                            help='Worksheet name to read (default: English). If not found, first sheet is used.')
        parser.add_argument('--upsert', action='store_true', help='Update existing rows (matched by errorcodenum)')
        parser.add_argument('--dry-run', action='store_true', help='Parse and validate but do not write to DB')

    def handle(self, *args, **options):
        file_path = options['file']
        sheet_name = options['sheet']
        do_upsert = options['upsert']
        dry_run = options['dry_run']

        # Ensure dependency is present
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.stdout.write(self.style.ERROR(
                'openpyxl is required. Install it with: pip install openpyxl'
            ))
            return

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        self.stdout.write(self.style.SUCCESS(f'Reading: {file_path}'))
        try:
            wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to open workbook: {e}'))
            return

        ws = None
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            if sheet_name and sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(
                    f'Sheet "{sheet_name}" not found. Using first sheet: {wb.sheetnames[0]}'
                ))
            ws = wb[wb.sheetnames[0]]

        # Find header row (first non-empty row)
        header_row_idx = None
        headers = []
        for r in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if r and any(c is not None and str(c).strip() != '' for c in r):
                headers = [str(c).strip() if c is not None else '' for c in r]
                header_row_idx = headers and 1 if ws.min_row == 1 else None
                break
        # If the above detection is not accurate, just assume first row is header
        if not headers:
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(c).strip() if c is not None else '' for c in first]
            header_row_idx = 1

        def norm(s: str) -> str:
            return (
                s.lower().strip()
                 .replace('\u200f', '')
                 .replace('\u200e', '')
                 .replace('\ufeff', '')
                 .replace('-', ' ')
                 .replace('_', ' ')
            )

        # Map possible header aliases to model fields
        aliases = {
            'errorcodenum': {
                'errorcodenum', 'error code', 'error code no', 'error code number', 'error no', 'error number', 'code', 'errcode', 'errorcode'
            },
            'errormessage': {
                'errormessage', 'error message', 'message', 'description', 'error description'
            },
            'probablecause': {
                'probablecause', 'probable cause', 'cause', 'reason'
            },
            'solution': {
                'solution', 'fix', 'resolution', 'action', 'how to fix', 'solutions'
            },
        }

        # Build column index map
        col_map = {}
        for idx, h in enumerate(headers):
            key = norm(h)
            for field, names in aliases.items():
                if key in names:
                    col_map[field] = idx
        if 'errorcodenum' not in col_map:
            self.stdout.write(self.style.ERROR(
                f"Could not locate 'error code' column. Headers detected: {headers}"
            ))
            return

        # Iterate rows
        import re
        def parse_code(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                try:
                    return int(val)
                except Exception:
                    return None
            s = str(val)
            m = re.search(r'\d+', s)
            return int(m.group(0)) if m else None

        def clean_text(val):
            if val is None:
                return ''
            return str(val).strip()

        created, updated, skipped, errors = 0, 0, 0, 0

        # Determine the first data row after header
        data_start = 2  # assume header at row 1 for read_only iterator

        # openpyxl read_only mode uses generator; use enumerate over ws.iter_rows
        with transaction.atomic():
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # extract values by map
                try:
                    code_val = row[col_map['errorcodenum']] if col_map['errorcodenum'] < len(row) else None
                    code = parse_code(code_val)
                    if code is None:
                        skipped += 1
                        continue

                    msg = clean_text(row[col_map['errormessage']]) if 'errormessage' in col_map and col_map['errormessage'] < len(row) else ''
                    cause = clean_text(row[col_map['probablecause']]) if 'probablecause' in col_map and col_map['probablecause'] < len(row) else ''
                    sol = clean_text(row[col_map['solution']]) if 'solution' in col_map and col_map['solution'] < len(row) else ''

                    if dry_run:
                        # Just validate existence/format
                        continue

                    if do_upsert:
                        obj, was_created = Errors.objects.get_or_create(
                            errorcodenum=code,
                            defaults={
                                'errormessage': msg,
                                'probablecause': cause,
                                'solution': sol,
                            }
                        )
                        if was_created:
                            created += 1
                        else:
                            # Update fields if changed
                            changed = False
                            if obj.errormessage != msg:
                                obj.errormessage = msg; changed = True
                            if obj.probablecause != cause:
                                obj.probablecause = cause; changed = True
                            if obj.solution != sol:
                                obj.solution = sol; changed = True
                            if changed:
                                obj.save(update_fields=['errormessage', 'probablecause', 'solution', 'updated_at'])
                                updated += 1
                    else:
                        # Insert only if not exists
                        if Errors.objects.filter(errorcodenum=code).exists():
                            skipped += 1
                        else:
                            Errors.objects.create(
                                errorcodenum=code,
                                errormessage=msg,
                                probablecause=cause,
                                solution=sol,
                            )
                            created += 1
                except Exception as e:
                    errors += 1
                    self.stdout.write(self.style.WARNING(f'Row {r_idx}: {e}'))

        if dry_run:
            self.stdout.write(self.style.SUCCESS('Dry-run completed successfully.'))
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('Import summary:'))
        self.stdout.write(f'  Created: {created}')
        self.stdout.write(f'  Updated: {updated}')
        self.stdout.write(f'  Skipped: {skipped}')
        self.stdout.write(f'  Errors:  {errors}')
